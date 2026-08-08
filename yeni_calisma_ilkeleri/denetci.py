#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DENETCI — G katmanı (işçilik): G01–G24 + G05b + G05c = 26 kapı.

    denetci.py     → G katmanı (işçilik)      ← BU BETİK
    olcek_testi    → Ö katmanı (dayanıklılık)
    deger_kapilari → D katmanı (değer)

Ölçtükleri: koruma (G01-G03), biçim (G04-G05c), ipucu/hata (G06-G08),
motor/karar yapısı (G09-G12), baskı (G13-G15), dil/kalite (G16-G20),
teknik sözleşme (G21-G24).

Kullanım:
    python denetci.py <dosya.xlsx> --spec SPEC.yaml [--rapor RAPOR_DENETIM.md]
"""

import sys, os, re, zipfile, argparse

try:
    import openpyxl
    from openpyxl.utils import range_boundaries
except ImportError:
    print("openpyxl gerekli:  pip install openpyxl"); sys.exit(2)

try:
    import yaml
except ImportError:
    yaml = None

SONUC = []
def kayit(k, s, m): SONUC.append((k, s, m))
def gec(k, m):  kayit(k, "GECTI", m)
def uyar(k, m): kayit(k, "UYARI", m)
def kal(k, m):  kayit(k, "KALDI", m)

HATA_DEGERLERI = ("#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#N/A", "#NUM!", "#NULL!", "#SPILL!")

# Ç1: yasak sözcük eşleşmesi kelime sınırıyla ve ≥4 harf (mandata §11)
YASAK_KELIMELER = ["lorem", "ipsum", "placeholder", "tbd", "dead", "stub",
                   "dummy", "fillme", "xxx", "todo", "fixme"]

# §11 kesin yasak fonksiyonlar (hata vermez, sessizce yanlış çalışır)
YASAK_FONKSIYONLAR = ["XLOOKUP", "XMATCH", "FILTER", "SORT", "SORTBY", "UNIQUE",
                      "SEQUENCE", "LAMBDA", "LET", "BYROW", "BYCOL", "MAP", "SCAN",
                      "REDUCE", "TEXTSPLIT", "VSTACK", "HSTACK", "TAKE", "DROP",
                      "GROUPBY", "PIVOTBY"]

# §11 _xlfn. öneki zorunlu fonksiyonlar
XLFN_ZORUNLU = ["TEXTJOIN", "CONCAT", "IFS", "SWITCH", "MAXIFS", "MINIFS"]

# bozuk UTF-8 / mojibake kalıpları (G17)
MOJIBAKE = re.compile(r"Ã[\x80-\xBF]|â€|\ufffd|â\x80|Ã\.")

def F(v): return isinstance(v, str) and v.startswith("=")
def metin(v): return "" if v is None else str(v)
def formuller(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if F(c.value):
                    yield ws, c


# ===========================================================================
def denetle(dosya, spec):
    wb = openpyxl.load_workbook(dosya)
    dm = spec["denetim"]

    # ---- G01-G03: koruma ----
    korumali_gerek = [dm["motor_sayfasi"], dm["karar_sayfasi"], dm["pano_sayfasi"], dm["ayarlar_sayfasi"]]
    korumasiz = [s for s in korumali_gerek if s not in wb.sheetnames]
    if korumasiz:
        kal("G01", f"Koruması beklenen sayfa yok: {korumasiz}")
    else:
        korumasiz = [s for s in korumali_gerek if not wb[s].protection.sheet]
        if korumasiz:
            kal("G01", f"Korumasız hesaplama sayfası: {korumasiz}")
        else:
            gec("G01", f"Hesaplama sayfaları korumalı: {korumali_gerek}")

    giris_sayfalari = [s["ad"] for s in spec["sayfalar"] if s.get("tur") == "giris"]
    kilitli_giris = [s for s in giris_sayfalari if wb[s].protection.sheet]
    if kilitli_giris:
        kal("G02", f"Giriş sayfaları kilitli — kullanıcı veri giremez: {kilitli_giris}")
    else:
        gec("G02", f"Giriş sayfaları veri girişine açık: {giris_sayfalari}")

    korumali_formul = sum(1 for s in korumali_gerek
                          if s in wb.sheetnames and any(F(c.value) for row in wb[s].iter_rows() for c in row))
    if korumali_formul == 0:
        kal("G03", "Korumalı sayfalarda hiç formül yok — koruma anlamsız")
    else:
        gec("G03", f"{korumali_formul}/{len([s for s in korumali_gerek if s in wb.sheetnames])} korumalı sayfada formül var")

    # ---- G04-G05c: biçim ----
    giris_dolgu = (dm.get("giris_dolgu") or ["FFFFF2CC"])[0].lstrip("#").upper()[-6:]
    ornek_huc = None
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=4, max_row=12, max_col=9):
            for c in row:
                if c.fill is not None and c.fill.start_color and c.fill.start_color.rgb:
                    rgb = str(c.fill.start_color.rgb).lstrip("#").upper()
                    if rgb.endswith(giris_dolgu):
                        ornek_huc = f"{ws.title}!{c.coordinate}"
                        break
                if ornek_huc: break
        if ornek_huc: break
    if not ornek_huc:
        kal("G04", f"Giriş alanı dolgu rengi {giris_dolgu} hiçbir örnek hücrede yok")
    else:
        gec("G04", f"Giriş dolgusu {giris_dolgu} → {ornek_huc}")

    baslik_sayfasi = None
    for ws in wb.worksheets:
        for c in ws[3]:
            if c.font is not None and c.font.bold and c.fill and c.fill.start_color and c.fill.start_color.rgb:
                baslik_sayfasi = ws.title
                break
        if baslik_sayfasi: break
    if not baslik_sayfasi:
        kal("G05", "Hiçbir tablo başlık satırı kalın + dolgulu değil")
    else:
        gec("G05", f"Başlık satırı kalın + dolgulu → {baslik_sayfasi}")

    formul_sayisi = sum(1 for _ in formuller(wb))
    kilitli_formul = sum(1 for ws, c in formuller(wb) if c.protection is None or c.protection.locked)
    if formul_sayisi == 0:
        kal("G05b", "Dosyada formül yok — çıktı katmanı üretilmemiş")
    elif kilitli_formul < max(1, int(formul_sayisi * 0.8)):
        kal("G05b", f"Formül hücrelerinin yalnızca {kilitli_formul}/{formul_sayisi} kilitli")
    else:
        gec("G05b", f"Formül hücreleri kilitli ({kilitli_formul}/{formul_sayisi})")

    sekme_rengi = sum(1 for ws in wb.worksheets
                      if ws.sheet_properties.tabColor is not None)
    if sekme_rengi == 0:
        kal("G05c", "Hiçbir sayfada sekme rengi yok")
    else:
        gec("G05c", f"{sekme_rengi} sayfada sekme rengi tanımlı")

    # ---- G06-G08: ipucu / hata / liste ----
    dv_ler = [dv for ws in wb.worksheets for dv in ws.data_validations.dataValidation]
    dv_prompt = [dv for dv in dv_ler if dv.promptTitle]
    dv_hata = [dv for dv in dv_ler if dv.errorTitle]
    if len(dv_ler) == 0:
        kal("G06", "Hiç veri doğrulama yok — giriş kılavuzu eksik")
    elif len(dv_prompt) < max(1, int(len(dv_ler) * 0.8)):
        kal("G06", f"Veri doğrulamaların {len(dv_prompt)}/{len(dv_ler)} ipucu (prompt) içeriyor")
    else:
        gec("G06", f"{len(dv_prompt)}/{len(dv_ler)} doğrulamada ipucu var")

    if len(dv_ler) == 0:
        kal("G07", "Hata mesajı denetlenemedi — veri doğrulama yok")
    elif len(dv_hata) < max(1, int(len(dv_ler) * 0.8)):
        kal("G07", f"Veri doğrulamaların {len(dv_hata)}/{len(dv_ler)} hata mesajı içeriyor")
    else:
        gec("G07", f"{len(dv_hata)}/{len(dv_ler)} doğrulamada hata mesajı var")

    liste_alanlar = [a for a in spec.get("alanlar", []) if a.get("tip") == "liste"]
    liste_dv = [dv for dv in dv_ler if dv.type == "list"]
    if liste_alanlar and not liste_dv:
        kal("G08", f"SPEC'te {len(liste_alanlar)} liste alanı var ama açılır liste (list DV) yok")
    elif liste_alanlar:
        gec("G08", f"{len(liste_dv)} açılır liste; {len(liste_alanlar)} liste alanı karşılanıyor")
    else:
        uyar("G08", "SPEC'te liste tipi alan tanımlı değil")

    # ---- G09-G12: motor / karar / pano / rapor ----
    mf = dm["motor_sayfasi"]
    motor_formul = sum(1 for ws, c in formuller(wb) if ws.title == mf)
    if motor_formul < 20:
        kal("G09", f"MOTOR'da yalnızca {motor_formul} formül hücresi")
    else:
        gec("G09", f"MOTOR'da {motor_formul} formül adımı")

    ksf = dm["karar_sayfasi"]
    if ksf not in wb.sheetnames:
        kal("G10", f"Karar sayfası yok: {ksf}")
    else:
        kf = sum(1 for row in wb[ksf].iter_rows() for c in row if F(c.value))
        sabit_karar = [c.coordinate for row in wb[ksf].iter_rows() for c in row
                       if c.value not in (None, "") and not F(c.value)][:5]
        if kf == 0 and sabit_karar:
            kal("G10", f"Karar hücreleri sabit metin, formül değil: {sabit_karar}")
        else:
            gec("G10", f"Karar sayfası formül türevli ({kf} formül)")

    pno = dm["pano_sayfasi"]
    if pno not in wb.sheetnames:
        kal("G11", f"PANO sayfası yok: {pno}")
    else:
        pano_formul = sum(1 for row in wb[pno].iter_rows() for c in row if F(c.value))
        if pano_formul < 10:
            kal("G11", f"PANO'da yalnızca {pano_formul} formül hücresi")
        else:
            gec("G11", f"PANO'da {pano_formul} canlı KPI formülü")

    rapor_sayfasi = next((s["ad"] for s in spec["sayfalar"] if s.get("tur") == "rapor"), None)
    if rapor_sayfasi:
        rf = sum(1 for row in wb[rapor_sayfasi].iter_rows() for c in row if F(c.value))
        if rf == 0:
            kal("G12", f"RAPOR sayfası formüle bağlı değil (0 formül)")
        else:
            gec("G12", f"RAPOR {rf} formülle bağlı")
    else:
        uyar("G12", "SPEC'te rapor sayfası tanımlı değil")

    # ---- G13-G15: baskı ----
    baski_sayfalari = [s["ad"] for s in spec["sayfalar"] if s.get("baski")]
    if not baski_sayfalari:
        uyar("G13", "SPEC'te baski:true sayfa yok")
    else:
        bos = [s for s in baski_sayfalari if not wb[s].print_area]
        if bos:
            kal("G13", f"Baskı alanı tanımsız: {bos}")
        else:
            gec("G13", f"Baskı alanı tanımlı: {baski_sayfalari}")

    bos_olcek = [s for s in baski_sayfalari
                 if not (wb[s].page_setup.fitToWidth or wb[s].page_setup.scale not in (None, 100))]
    if baski_sayfalari and bos_olcek:
        kal("G14", f"Sayfa genişliği/ölçek ayarı yok: {bos_olcek}")
    elif baski_sayfalari:
        gec("G14", f"Yazdırma ölçeği ayarlı: {baski_sayfalari}")

    orta = [s for s in baski_sayfalari if not (wb[s].print_options.horizontalCentered or wb[s].print_title_rows)]
    if baski_sayfalari and orta:
        kal("G15", f"Yazdırma düzeni (yatay ortala / başlık satırı) yok: {orta}")
    elif baski_sayfalari:
        gec("G15", "Baskı düzeni ayarlı")

    # ---- G16-G20: dil / güvenlik ----
    yasak_bulgu = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                for w in re.findall(r"[A-Za-zÇĞİÖŞÜçğıöşü]{4,}", metin(c.value)):
                    if w.lower() in YASAK_KELIMELER:
                        yasak_bulgu.append(f"{ws.title}!{c.coordinate}={w}")
    if yasak_bulgu:
        kal("G16", f"Yasak sözcük: {yasak_bulgu[:5]}")
    else:
        gec("G16", "Yasak sözcük yok")

    bozuk = [f"{ws.title}!{c.coordinate}" for ws, c in formuller(wb)
             if MOJIBAKE.search(metin(c.value))]
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None and MOJIBAKE.search(metin(c.value)):
                    bozuk.append(f"{ws.title}!{c.coordinate}")
    if bozuk:
        kal("G17", f"Bozuk Türkçe karakter (mojibake): {bozuk[:5]}")
    else:
        gec("G17", "Türkçe karakterler temiz")

    yasak_fx = []
    for ws, c in formuller(wb):
        ust = metin(c.value).upper()
        for fx in YASAK_FONKSIYONLAR:
            if re.search(rf"\b{fx}\(", ust) and f"_XLFN.{fx}" not in ust:
                yasak_fx.append(f"{ws.title}!{c.coordinate}={fx}")
                break
    if yasak_fx:
        kal("G18", f"Yasak fonksiyon: {yasak_fx[:5]}")
    else:
        gec("G18", "Yasak fonksiyon kullanılmamış")

    cplak_xlfn = []
    for ws, c in formuller(wb):
        for fx in XLFN_ZORUNLU:
            if re.search(rf"(?<!_xlfn\.)\b{fx}\(", metin(c.value), re.IGNORECASE):
                cplak_xlfn.append(f"{ws.title}!{c.coordinate}={fx}")
                break
    if cplak_xlfn:
        kal("G19", f"_xlfn. öneki eksik: {cplak_xlfn[:5]}")
    else:
        gec("G19", "_xlfn. önekleri tutarlı")

    makro = "vbaProject.bin" in (zipfile.ZipFile(dosya).namelist() if zipfile.is_zipfile(dosya) else [])
    if makro:
        kal("G20", "Dosya makro (VBA) içeriyor — pasif kullanım güvenliği ihlali")
    else:
        gec("G20", "Makro yok")

    # ---- G21-G24: teknik sözleşme ----
    kb_sayisi = sum(len(rng.rules) for ws in wb.worksheets for rng in ws.conditional_formatting)
    asgari_kb = int(dm.get("asgari_kb_kurali_toplam", 30))
    if kb_sayisi < asgari_kb:
        kal("G21", f"Koşullu biçimlendirme {kb_sayisi} (asgari {asgari_kb})")
    else:
        gec("G21", f"{kb_sayisi} koşullu biçimlendirme kuralı")

    adlar = dict(wb.defined_names)
    gecersiz = []
    for ad, tanim in adlar.items():
        try:
            hedef = None
            for wsname, cellref in tanim.destinations:
                hedef = (wsname, cellref)
                break
            if hedef is None:
                continue
            ws, ref = hedef
            if ws in wb.sheetnames:
                wb[ws][ref]
        except Exception:
            gecersiz.append(ad)
    if gecersiz:
        kal("G22", f"Çözümlenemeyen ad tanımı: {gecersiz[:5]}")
    else:
        gec("G22", f"{len(adlar)} ad tanımı çözümleniyor")

    wb2 = openpyxl.load_workbook(dosya, data_only=True)
    hata_hucre = []
    for ws in wb2.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if metin(c.value).strip() in HATA_DEGERLERI:
                    hata_hucre.append(f"{ws.title}!{c.coordinate}={c.value}")
    if hata_hucre:
        kal("G23", f"Dosyada formül hatası: {hata_hucre[:5]}")
    else:
        gec("G23", "Formül hataları yok")

    asgari_kap = int(dm.get("asgari_tablo_kapasitesi", 1000))
    kucuk = []
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            mn, mr, mx, sr = range_boundaries(tbl.ref)
            kap = sr - mr
            if kap < asgari_kap:
                kucuk.append(f"{tbl.name}={kap}")
    if kucuk:
        kal("G24", f"Tablo kapasitesi < {asgari_kap}: {kucuk}")
    else:
        gec("G24", f"Tüm tablo kapasiteleri ≥ {asgari_kap}")


# ===========================================================================
def rapor_yaz(yol, dosya, pi):
    kaldi = [s for s in SONUC if s[1] == "KALDI"]
    uyari = [s for s in SONUC if s[1] == "UYARI"]
    gecti = [s for s in SONUC if s[1] == "GECTI"]
    sat = ["# DENETİM RAPORU (G01–G24)", "", f"Dosya: `{dosya}`", f"SHA-256: `{pi}`", "",
           f"- GEÇTİ: {len(gecti)}", f"- UYARI: {len(uyari)}", f"- **KALDI: {len(kaldi)}**", "", "## Kalan kapılar", ""]
    for k, s, m in kaldi: sat.append(f"- **{k}** — {m}")
    sat += ["", "## Uyarılar", ""]
    for k, s, m in uyari: sat.append(f"- {k} — {m}")
    sat += ["", "## Geçen kapılar", ""]
    for k, s, m in gecti: sat.append(f"- {k} — {m}")
    with open(yol, "w", encoding="utf-8") as f:
        f.write("\n".join(sat))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("dosya")
    ap.add_argument("--spec", default=None)
    ap.add_argument("--rapor", default="RAPOR_DENETIM.md")
    a = ap.parse_args()

    spec = {}
    if a.spec:
        if yaml is None:
            print("[uyari] pyyaml yok — SPEC parametreleri kullanılamıyor"); sys.exit(2)
        with open(a.spec, encoding="utf-8") as f:
            spec = yaml.safe_load(f) or {}

    import hashlib
    pi = hashlib.sha256(open(a.dosya, "rb").read()).hexdigest()

    denetle(a.dosya, spec)
    rapor_yaz(a.rapor, a.dosya, pi)

    kaldi = [s for s in SONUC if s[1] == "KALDI"]
    for k, s, m in SONUC:
        isaret = {"GECTI": "  GEÇTİ", "UYARI": "  UYARI", "KALDI": "! KALDI"}[s]
        print(f"{isaret}  {k:5s} {m}")
    print("\n" + "=" * 74)
    print(f"SONUÇ: {len(kaldi)} KALDI | rapor: {a.rapor}")
    print("SEVK EDİLEMEZ" if kaldi else "SEVK EDİLEBİLİR")
    sys.exit(1 if kaldi else 0)


if __name__ == "__main__":
    main()
