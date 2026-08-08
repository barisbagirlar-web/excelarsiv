#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DEGER_KAPILARI — v4 değer katmanı doğrulayıcısı  (D01..D14)

denetci.py (G01..G24) işçiliği ölçer. Bu betik DEĞERİ ölçer:
ölü çıktı, sahte KPI, boş dosyada yanlış karar, çapraz tutarsızlık,
analitik derinlik, kaynak katmanı.

Kullanım:
    python deger_kapilari.py <dosya.xlsx> [--spec SPEC.yaml] [--rapor RAPOR_DEGER.md]
    python deger_kapilari.py <dosya.xlsx> --calisma-zamani     # D11/D12 için LibreOffice ile

Tek KALDI varsa çıkış kodu 1'dir.
Hiçbir kapı beyana güvenmez; hepsi dosyadan ölçülür.
"""

import sys, re, json, shutil, zipfile, argparse, subprocess, tempfile, hashlib, os
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.utils import range_boundaries
except ImportError:
    print("openpyxl gerekli:  pip install openpyxl"); sys.exit(2)

# ---------------------------------------------------------------------------
VARSAYILAN = {
    "asgari_grafik": 8,
    "asgari_pano_kpi": 12,
    "asgari_motor_adimi": 40,
    "asgari_derinlik_puani": 100,
    "asgari_ileri_modul": 2,
    "ikame_carpani": 3.0,
    "asgari_ayrim_maddesi": 5,
    "capraz_tutarlilik_tolerans": 0.01,
    "bos_dosya_beklenen_karar": "VERİ YOK",
    "karar_sayfasi": "KARAR",
    "motor_sayfasi": "MOTOR",
    "pano_sayfasi": "PANO",
    "ayarlar_sayfasi": "AYARLAR",
    "modul_puanlari": {"T": 10, "O": 15, "I": 25},
    "yorum_fonksiyonlari": ["TEXTJOIN", "CONCAT", "CONCATENATE", "&"],
}

SONUC = []
def kayit(k, s, m): SONUC.append((k, s, m))
def gec(k, m):  kayit(k, "GECTI", m)
def uyar(k, m): kayit(k, "UYARI", m)
def kal(k, m):  kayit(k, "KALDI", m)

def F(v):  # formül mü
    return isinstance(v, str) and v.startswith("=")

def metin(v):
    return "" if v is None else str(v)

def sayisal(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


# ===========================================================================
def statik_denetim(yol, spec):
    wb  = openpyxl.load_workbook(yol, data_only=False)
    wbd = openpyxl.load_workbook(yol, data_only=True)

    sayfalar = wb.sheetnames
    adlar = {k: v.value for k, v in wb.defined_names.items()}

    # tüm formülleri topla
    formuller = []   # (sayfa, koord, formul)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if F(c.value):
                    formuller.append((ws.title, c.coordinate, c.value))

    # tabloları topla
    tablolar = {}
    for ws in wb.worksheets:
        for t in ws.tables.values():
            tablolar[(ws.title, t.name)] = t

    # ---- D14  FORMÜL AYRACI ------------------------------------------------
    # XML formülü daima virgül ister. ';' varsa dosya açılışta bozulur.
    ayrac = []
    for sn, co, f in formuller:
        govde = re.sub(r'"[^"]*"', '', f)          # string literalleri çıkar
        if ";" in govde:
            ayrac.append((sn, co))
    if ayrac:
        kal("D14", f"{len(ayrac)} formülde ';' ayracı — Excel dosyayı bozuk açar: {ayrac[:5]}")
    else:
        gec("D14", f"{len(formuller)} formülde ayraç doğru (virgül)")

    # ---- D07  ÖLÜ ÇIKTI  (en pahalı kusur) ---------------------------------
    olu = []
    for (sn, tn), t in tablolar.items():
        ws = wb[sn]
        mn, mr, mx, sr = range_boundaries(t.ref)
        basliklar = [ws.cell(mr, c).value for c in range(mn, mx + 1)]
        # dolu satır aralığını bul
        dolu_satirlar = []
        for r in range(mr + 1, sr + 1):
            if any(ws.cell(r, c).value not in (None, "") for c in range(mn, mx + 1)):
                dolu_satirlar.append(r)
        if not dolu_satirlar:
            continue
        for ci, c in enumerate(range(mn, mx + 1)):
            basl = metin(basliklar[ci]).strip()
            if not basl:
                continue
            degerler = [ws.cell(r, c).value for r in dolu_satirlar]
            formul_var = any(F(v) for v in degerler)
            deger_var  = any(v not in (None, "") for v in degerler)
            if not formul_var and not deger_var:
                olu.append(f"{sn}.{tn}[{basl}] — başlık var, {len(dolu_satirlar)} dolu satırda ne formül ne değer")
    if olu:
        kal("D07", f"ÖLÜ ÇIKTI — {len(olu)} kolon hesaplanmıyor: {olu[:6]}")
    else:
        gec("D07", "Ölü çıktı kolonu yok")

    # ---- D07b  ÖLÜ AD TANIMI ----------------------------------------------
    olu_ad = []
    for ad, ref in adlar.items():
        m = re.match(r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)$", metin(ref))
        if not m:
            continue
        sn, col, row = m.group(1), m.group(2), int(m.group(3))
        if sn not in wb.sheetnames:
            continue
        h = wb[sn][f"{col}{row}"]
        if h.value in (None, ""):
            olu_ad.append(f"{ad} → {sn}!{col}{row} BOŞ")
    if olu_ad:
        kal("D07b", f"{len(olu_ad)} ad tanımı boş hücreyi gösteriyor: {olu_ad[:5]}")
    else:
        gec("D07b", f"{len(adlar)} ad tanımının hedefi dolu")

    # ---- D06  SAHTE KPI ----------------------------------------------------
    # Bir KPI yalnızca sabit sayılara bağlıysa veriyi hiç ölçmüyordur.
    ayarlar_sabitleri = set()
    asf = spec["ayarlar_sayfasi"]
    if asf in wb.sheetnames:
        for ad, ref in adlar.items():
            if asf in metin(ref):
                m = re.search(r"\$?([A-Z]+)\$?(\d+)$", metin(ref))
                if m:
                    h = wb[asf][f"{m.group(1)}{m.group(2)}"]
                    if not F(h.value) and sayisal(h.value):
                        ayarlar_sabitleri.add(ad)

    tablo_adlari = {tn for (_, tn) in tablolar}
    sahte = []
    for sn, co, f in formuller:
        if sn in (asf,):
            continue
        # tabloya veya başka sayfa aralığına hiç dokunmuyor mu?
        tabloya_dokunuyor = any(tn in f for tn in tablo_adlari)
        govde_ = re.sub(r'"[^"]*"', '', f)
        araliga_dokunuyor = bool(re.search(r"\$?[A-Z]{1,3}\$?\d+", govde_))
        kullanilan_adlar  = [a for a in adlar if re.search(rf"\b{re.escape(a)}\b", f)]
        sadece_sabit = (kullanilan_adlar
                        and all(a in ayarlar_sabitleri for a in kullanilan_adlar))
        if sadece_sabit and not tabloya_dokunuyor and not araliga_dokunuyor:
            sahte.append((sn, co, f[:60]))
    if sahte:
        kal("D06", f"SAHTE KPI — {len(sahte)} hücre yalnızca sabitlerden türüyor, veriyi ölçmüyor: {sahte[:5]}")
    else:
        gec("D06", "Sabitten türeyen sahte KPI yok")

    # ---- D13  MOTOR DERİNLİĞİ ---------------------------------------------
    motor_adim = 0
    msf = spec["motor_sayfasi"]
    motor_sayfalari = [s for s in sayfalar if msf in s.upper() or "HESAP" in s.upper()]
    for s in motor_sayfalari:
        motor_adim += sum(1 for r in wb[s].iter_rows() for c in r if F(c.value))
    if motor_adim < spec["asgari_motor_adimi"]:
        kal("D13", f"Motor adımı {motor_adim} (asgari {spec['asgari_motor_adimi']}) — "
                   f"hesap katmanı sığ. Sayfalar: {motor_sayfalari}")
    else:
        gec("D13", f"Motor {motor_adim} isimli/hesaplı adım")

    # ---- D08  GÖRÜNÜR DERİNLİK (PANO KPI) ---------------------------------
    psf = spec["pano_sayfasi"]
    pano = [s for s in sayfalar if psf in s.upper() or "DASHBOARD" in s.upper()]
    if not pano:
        kal("D08", "PANO sayfası yok")
    else:
        kpi = set()
        for s in pano:
            for r in wb[s].iter_rows():
                for c in r:
                    if F(c.value):
                        kpi.add(c.value.strip())
        if len(kpi) < spec["asgari_pano_kpi"]:
            kal("D08", f"PANO'da farklı KPI {len(kpi)} (asgari {spec['asgari_pano_kpi']})")
        else:
            gec("D08", f"PANO'da {len(kpi)} farklı canlı KPI")

    # ---- D09  GRAFİK + TAHMİN ARALIĞI -------------------------------------
    gsay = sum(len(ws._charts) for ws in wb.worksheets)
    tahmin_var = any(re.search(r"FORECAST|TREND|SLOPE|PERCENTILE", f.upper()) for _, _, f in formuller)
    if gsay < spec["asgari_grafik"]:
        kal("D09", f"Grafik {gsay} (asgari {spec['asgari_grafik']})")
    elif not tahmin_var:
        kal("D09", f"{gsay} grafik var ama tahmin/aralık hesabı yok (FORECAST/TREND/PERCENTILE)")
    else:
        gec("D09", f"{gsay} grafik + tahmin katmanı mevcut")

    # ---- D09b BOŞ GRAFİK ---------------------------------------------------
    bos_grafik = []
    for ws in wb.worksheets:
        for ch in ws._charts:
            for ser in ch.series:
                ref = None
                if ser.val is not None and ser.val.numRef is not None:
                    ref = ser.val.numRef.f
                if not ref:
                    continue
                m = re.match(r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)$", ref)
                if not m:
                    continue
                sn = m.group(1)
                if sn not in wbd.sheetnames:
                    continue
                w = wbd[sn]
                vals = []
                for rr in range(int(m.group(3)), int(m.group(5)) + 1):
                    vals.append(w[f"{m.group(2)}{rr}"].value)
                sayilar = [v for v in vals if sayisal(v)]
                if not sayilar or all(v == 0 for v in sayilar):
                    bos_grafik.append(f"{ws.title}: {ref} tamamı sıfır/boş")
    if bos_grafik:
        kal("D09b", f"{len(bos_grafik)} grafik serisi düz sıfır — alıcı ilk ekranda boş görür: {bos_grafik[:5]}")
    else:
        gec("D09b", "Grafik serilerinde veri var")

    # ---- D10  KAYNAK KATMANI ----------------------------------------------
    if asf not in wb.sheetnames:
        kal("D10", f"{asf} sayfası yok")
    else:
        ws = wb[asf]
        basl_satir, kolonlar = None, {}
        for r in range(1, min(ws.max_row, 30) + 1):
            satir = [metin(c.value).strip().lower() for c in ws[r]]
            if "anahtar" in satir:
                basl_satir = r
                for i, v in enumerate(satir, start=1):
                    kolonlar[v] = i
                break
        if basl_satir is None:
            kal("D10", f"{asf}: 'anahtar' başlıklı parametre tablosu bulunamadı")
        else:
            eksik_kolon = [k for k in ("kaynak", "yururluk_tarihi") if k not in kolonlar]
            if eksik_kolon:
                kal("D10", f"{asf}: zorunlu kolon yok: {eksik_kolon} | mevcut: {list(kolonlar)}")
            else:
                bos = []
                for r in range(basl_satir + 1, ws.max_row + 1):
                    if ws.cell(r, kolonlar["anahtar"]).value in (None, ""):
                        continue
                    for k in ("kaynak", "yururluk_tarihi"):
                        if ws.cell(r, kolonlar[k]).value in (None, ""):
                            bos.append(f"satır {r} / {k}")
                if bos:
                    kal("D10", f"{len(bos)} parametrede kaynak/yürürlük boş: {bos[:6]}")
                else:
                    gec("D10", "Tüm parametrelerde kaynak ve yürürlük tarihi dolu")

    # ---- D12  ÇAPRAZ TUTARLILIK -------------------------------------------
    etiket_deger = defaultdict(list)
    haric = ("ORNEK", "ÖRNEK", "DEMO", "TEST", "LISTELER", "DEGISIKLIK")
    for ws in wbd.worksheets:
        if any(h in ws.title.upper() for h in haric):
            continue
        for row in ws.iter_rows():
            hucreler = [c for c in row if c.value not in (None, "")]
            for i in range(len(hucreler) - 1):
                et, dg = hucreler[i].value, hucreler[i + 1].value
                if isinstance(et, str) and len(et.strip()) > 8 and sayisal(dg):
                    etiket_deger[et.strip().lower()].append((ws.title, hucreler[i + 1].coordinate, dg))
    catisma = []
    tol = spec["capraz_tutarlilik_tolerans"]
    for et, kayitlar in etiket_deger.items():
        if len(kayitlar) < 2:
            continue
        degerler = [k[2] for k in kayitlar]
        enb, enk = max(degerler), min(degerler)
        if enb == 0 and enk == 0:
            continue
        taban = max(abs(enb), abs(enk)) or 1
        if abs(enb - enk) / taban > tol:
            catisma.append(f"'{et}' → {[(k[0], k[1], k[2]) for k in kayitlar]}")
    if catisma:
        kal("D12", f"ÇAPRAZ TUTARSIZLIK — aynı etiket farklı değer: {catisma[:4]}")
    else:
        gec("D12", "Çapraz tutarlılık sağlandı")

    # ---- D03/D04/D05  ANALİTİK DERİNLİK (spec beyanı + dosya karşılığı) ----
    moduller = spec.get("analitik_moduller") or []
    if not moduller:
        uyar("D03", "SPEC'te analitik_moduller yok — derinlik puanı ölçülemedi")
    else:
        puan, ileri, eksik, statik_yorum = 0, 0, [], []
        for m in moduller:
            kod = metin(m.get("kod")).strip().upper()
            at  = metin(m.get("ad_tanimi")).strip()
            yh  = metin(m.get("yorum_hucresi")).strip()
            if at and at not in adlar:
                eksik.append(f"{kod}: ad tanımı '{at}' dosyada yok"); continue
            # yorum hücresi canlı mı
            if yh:
                if yh not in adlar:
                    eksik.append(f"{kod}: yorum hücresi '{yh}' yok"); continue
                r = metin(adlar[yh])
                mm = re.match(r"^'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)$", r)
                if mm and mm.group(1) in wb.sheetnames:
                    hv = wb[mm.group(1)][f"{mm.group(2)}{mm.group(3)}"].value
                    if not F(hv):
                        statik_yorum.append(f"{kod}: yorum sabit metin, makine üretmiyor"); continue
            puan += spec["modul_puanlari"].get(kod[:1], 0)
            if kod.startswith("I"):
                ileri += 1
        if eksik:
            kal("D03", f"Beyan edilen modülün dosyada karşılığı yok: {eksik[:5]}")
        elif puan < spec["asgari_derinlik_puani"]:
            kal("D03", f"Derinlik puanı {puan} (asgari {spec['asgari_derinlik_puani']})")
        else:
            gec("D03", f"Derinlik puanı {puan}")
        if statik_yorum:
            kal("D05", f"Yorum cümlesi sabit metin: {statik_yorum[:5]}")
        else:
            gec("D05", "Modül yorumları makine üretimli")
        if ileri < spec["asgari_ileri_modul"]:
            kal("D04", f"İleri modül {ileri} (asgari {spec['asgari_ileri_modul']})")
        else:
            gec("D04", f"{ileri} ileri modül")

    # ---- D01  İKAME MALİYETİ ÇAPASI ---------------------------------------
    d = spec.get("deger") or {}
    if not d or not any(d.values()):
        uyar("D01", "SPEC'te deger bloğu doldurulmadı — fiyat çapası ölçülemedi (sahibi dolduracak)")
    else:
        fiyat = float(d.get("satis_fiyati_tl", 0) or 0)
        y1 = float(d.get("Y1_danisman_ikamesi_tl", 0) or 0)
        y2 = float(d.get("Y2_kurulum_ikamesi_tl", 0) or 0)
        esik = spec["ikame_carpani"] * fiyat
        if fiyat <= 0:
            kal("D01", "satis_fiyati_tl tanımsız")
        elif (y1 + y2) < esik:
            kal("D01", f"İkame maliyeti {y1+y2:,.0f} TL < eşik {esik:,.0f} TL "
                       f"({spec['ikame_carpani']}× {fiyat:,.0f}) — bu fiyat savunulamaz")
        else:
            gec("D01", f"İkame {y1+y2:,.0f} TL ≥ eşik {esik:,.0f} TL")

    # ---- D02  SERBEST ALTERNATİF AYRIMI ------------------------------------
    sa = spec.get("serbest_alternatif") or {}
    ayrimlar = sa.get("ayrimlar") or []
    if not sa or not ayrimlar:
        uyar("D02", "SPEC'te serbest_alternatif doldurulmadı — ayrım ölçülemedi (sahibi dolduracak)")
    elif len(ayrimlar) < spec["asgari_ayrim_maddesi"]:
        kal("D02", f"Ayrım maddesi {len(ayrimlar)} (asgari {spec['asgari_ayrim_maddesi']})")
    else:
        karsiliksiz = [a.get("madde") for a in ayrimlar
                       if metin(a.get("karsilik")).strip() not in adlar]
        if karsiliksiz:
            kal("D02", f"Ayrım maddesinin dosyada karşılığı yok: {karsiliksiz[:4]}")
        else:
            gec("D02", f"{len(ayrimlar)} ayrım maddesi, hepsi dosyada karşılıklı")

    return wb, wbd, tablolar


# ===========================================================================
def recalc(yol, timeout=180):
    """LibreOffice ile yeniden hesapla (yerinde)."""
    betik = os.path.join(os.path.dirname(os.path.abspath(__file__)), "recalc.py")
    if not os.path.exists(betik):
        return None
    try:
        r = subprocess.run([sys.executable, betik, yol, str(timeout)],
                           capture_output=True, text=True, timeout=timeout + 60)
        return json.loads(r.stdout) if r.stdout.strip().startswith("{") else None
    except Exception:
        return None


def calisma_zamani_denetim(yol, spec, tablolar):
    """D11 — veriyi boşalt, karar hücresi ne diyor?"""
    ksf = spec["karar_sayfasi"]
    with tempfile.TemporaryDirectory() as td:
        kopya = os.path.join(td, "bos.xlsx")
        shutil.copy(yol, kopya)
        wb = openpyxl.load_workbook(kopya)
        if ksf not in wb.sheetnames:
            kal("D11", f"{ksf} sayfası yok — karar katmanı ölçülemedi"); return
        silinen = 0
        for (sn, tn), t in tablolar.items():
            ws = wb[sn]
            mn, mr, mx, sr = range_boundaries(t.ref)
            for r in range(mr + 1, sr + 1):
                for c in range(mn, mx + 1):
                    cell = ws.cell(r, c)
                    if not F(cell.value) and cell.value not in (None, ""):
                        cell.value = None; silinen += 1
        if silinen == 0:
            uyar("D11", "Boşaltılacak veri bulunamadı"); return
        wb.save(kopya)
        sonuc = recalc(kopya)
        if sonuc is None:
            uyar("D11", "LibreOffice ile yeniden hesaplama yapılamadı"); return
        wbd = openpyxl.load_workbook(kopya, data_only=True)
        ws = wbd[ksf]
        karar = None
        for row in ws.iter_rows():
            for i, c in enumerate(row):
                if metin(c.value).strip().upper() == "KARAR" and i + 1 < len(row):
                    karar = row[i + 1].value
            if karar is not None:
                break
        beklenen = spec["bos_dosya_beklenen_karar"]
        if karar is None:
            uyar("D11", "Karar hücresi bulunamadı")
        elif metin(karar).strip().upper() == metin(beklenen).strip().upper():
            gec("D11", f"Boş dosyada karar = '{karar}' (doğru)")
        else:
            kal("D11", f"BOŞ DOSYADA YANLIŞ KARAR — {silinen} veri silindi, "
                       f"karar hâlâ '{karar}' (beklenen '{beklenen}'). "
                       f"Sessiz yanlış-pozitif: dosya veri yokluğunu 'her şey yolunda' sanıyor.")


# ===========================================================================
def parmak_izi(yol):
    h = hashlib.sha256()
    with open(yol, "rb") as f:
        for blok in iter(lambda: f.read(65536), b""):
            h.update(blok)
    return h.hexdigest()


def rapor_yaz(yol, dosya, pi):
    kaldi = [s for s in SONUC if s[1] == "KALDI"]
    uyari = [s for s in SONUC if s[1] == "UYARI"]
    gecti = [s for s in SONUC if s[1] == "GECTI"]
    sat = ["# DEĞER KAPILARI RAPORU (D01–D14)", "",
           f"Dosya: `{dosya}`", f"SHA-256: `{pi}`", "",
           f"- GEÇTİ: {len(gecti)}", f"- UYARI: {len(uyari)}",
           f"- **KALDI: {len(kaldi)}**", "", "## Kalan kapılar", ""]
    for k, s, m in kaldi: sat.append(f"- **{k}** — {m}")
    sat += ["", "## Uyarılar", ""]
    for k, s, m in uyari: sat.append(f"- {k} — {m}")
    sat += ["", "## Geçen kapılar", ""]
    for k, s, m in gecti: sat.append(f"- {k} — {m}")
    with open(yol, "w", encoding="utf-8") as f:
        f.write("\n".join(sat))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("dosya")
    ap.add_argument("--spec", default=None)
    ap.add_argument("--rapor", default="RAPOR_DEGER.md")
    ap.add_argument("--calisma-zamani", action="store_true",
                    help="D11 için LibreOffice ile boş-dosya testi çalıştır")
    a = ap.parse_args()

    spec = dict(VARSAYILAN)
    if a.spec:
        try:
            import yaml
            with open(a.spec, encoding="utf-8") as f:
                y = yaml.safe_load(f) or {}
            spec.update(y.get("denetim", {}))
            for blok in ("deger", "serbest_alternatif", "analitik_moduller"):
                if blok in y:
                    spec[blok] = y[blok]
        except Exception as e:
            print(f"[uyari] spec okunamadi: {e}")

    wb, wbd, tablolar = statik_denetim(a.dosya, spec)
    if a.calisma_zamani:
        calisma_zamani_denetim(a.dosya, spec, tablolar)

    pi = parmak_izi(a.dosya)
    rapor_yaz(a.rapor, a.dosya, pi)

    kaldi = [s for s in SONUC if s[1] == "KALDI"]
    for k, s, m in SONUC:
        isaret = {"GECTI": "  GEÇTİ", "UYARI": "  UYARI", "KALDI": "! KALDI"}[s]
        print(f"{isaret}  {k:6s} {m}")
    print("\n" + "=" * 74)
    print(f"SHA-256: {pi}")
    print(f"SONUÇ: {len(kaldi)} KALDI | rapor: {a.rapor}")
    print("SEVK EDİLEMEZ" if kaldi else "SEVK EDİLEBİLİR")
    sys.exit(1 if kaldi else 0)


if __name__ == "__main__":
    main()
