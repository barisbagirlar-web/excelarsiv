#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
OLCEK_TESTI — Ö1 (ölçek) ve Ö2 (uç durum) kapıları

denetci.py     → G katmanı (işçilik)
deger_kapilari → D katmanı (değer)
olcek_testi    → Ö katmanı (dayanıklılık)   ← BU BETİK

Ö1: Tabloya N satır enjekte eder, hesaplanan kolon formüllerini aşağı taşır,
    LibreOffice ile yeniden hesaplar, SÜREYİ ölçer ve formül hatası arar.
Ö2: Yedi uç durum senaryosunu tek tek uygular; her birinde
    (a) kullanıcıya görünen hata değeri (#DIV/0! vb.) çıkmamalı,
    (b) karar katmanı sessizce "UYGUN" dememeli.

Kullanım:
    python olcek_testi.py <dosya.xlsx> [--spec SPEC.yaml] [--satir 5000]
                          [--sure 5] [--rapor RAPOR_OLCEK.md] [--senaryo tumu_bos]

Tek KALDI varsa çıkış kodu 1'dir.
"""

import sys, os, re, json, time, shutil, argparse, tempfile, subprocess, random
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.utils import range_boundaries, get_column_letter
    from openpyxl.formula.translate import Translator
    from openpyxl.cell.cell import MergedCell
except ImportError:
    print("openpyxl gerekli:  pip install openpyxl"); sys.exit(2)

HATA_DEGERLERI = ("#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#N/A", "#NUM!", "#NULL!", "#SPILL!")

VARSAYILAN = {
    "olcek_satir": 5000,
    "azami_hesap_saniye": 5.0,
    "karar_sayfasi": "KARAR",
    "bos_dosya_beklenen_karar": "VERİ YOK",
    "asgari_dolu_satir": 3,
}

SONUC = []
def kayit(k, s, m): SONUC.append((k, s, m))
def gec(k, m):  kayit(k, "GECTI", m)
def uyar(k, m): kayit(k, "UYARI", m)
def kal(k, m):  kayit(k, "KALDI", m)

def F(v): return isinstance(v, str) and v.startswith("=")
def metin(v): return "" if v is None else str(v)
def sayisal(v): return isinstance(v, (int, float)) and not isinstance(v, bool)


# ---------------------------------------------------------------------------
def recalc(yol, timeout=300):
    """LibreOffice ile yeniden hesapla. (sonuc_dict, gecen_saniye) döner."""
    betik = os.path.join(os.path.dirname(os.path.abspath(__file__)), "recalc.py")
    if not os.path.exists(betik):
        return None, 0.0
    t0 = time.time()
    try:
        r = subprocess.run([sys.executable, betik, yol, str(timeout)],
                           capture_output=True, text=True, timeout=timeout + 120)
        gecen = time.time() - t0
        cikti = r.stdout.strip()
        i = cikti.find("{")
        return (json.loads(cikti[i:]) if i >= 0 else None), gecen
    except Exception as e:
        return {"error": str(e)}, time.time() - t0


def tablolari_bul(wb):
    t = {}
    for ws in wb.worksheets:
        for tbl in ws.tables.values():
            t[(ws.title, tbl.name)] = tbl
    return t


def kolon_profili(ws, mn, mr, sr, mx):
    """Her kolon için: (tip, formul_sablonu, kaynak_satir, ornek_degerler)"""
    profil = {}
    dolu = [r for r in range(mr + 1, sr + 1)
            if any(ws.cell(r, c).value not in (None, "") for c in range(mn, mx + 1))]
    if not dolu:
        return profil, []
    for c in range(mn, mx + 1):
        formul, kaynak = None, None
        ornekler = []
        for r in dolu:
            v = ws.cell(r, c).value
            if F(v) and formul is None:
                formul, kaynak = v, r
            elif v not in (None, ""):
                ornekler.append(v)
        if formul:
            tip = "formul"
        elif not ornekler:
            tip = "bos"
        elif all(isinstance(v, datetime) for v in ornekler):
            tip = "tarih"
        elif all(sayisal(v) for v in ornekler):
            tip = "sayi"
        else:
            tip = "metin"
        profil[c] = {"tip": tip, "formul": formul, "kaynak": kaynak, "ornekler": ornekler}
    return profil, dolu


def deger_uret(prof, i, senaryo):
    """Senaryoya göre tek hücre değeri üret."""
    tip = prof["tip"]
    orn = prof["ornekler"]
    if tip == "bos":
        return None
    if senaryo == "tumu_bos":
        return None
    if tip == "tarih":
        taban = orn[0] if orn else datetime(2026, 1, 1)
        if senaryo == "gelecek_tarih":
            return taban + timedelta(days=3000)
        return taban + timedelta(days=i)
    if tip == "sayi":
        taban = float(orn[0]) if orn else 1000.0
        if senaryo == "tumu_sifir":   return 0
        if senaryo == "negatif_deger": return -abs(taban)
        if senaryo == "azami_deger":   return 1e12
        if senaryo == "yinelenen_kayit": return taban
        return round(taban * random.uniform(0.6, 1.4), 2)
    # metin
    taban = metin(orn[0]) if orn else "Kayit"
    if senaryo == "yinelenen_kayit":
        return taban
    return f"{taban[:20]} {i}"


def varyant_uret(kaynak_yol, hedef_yol, hedef_satir, senaryo, tek_satir=False):
    """Tabloları hedef_satir'a kadar doldur; formülleri aşağı taşı."""
    shutil.copy(kaynak_yol, hedef_yol)
    wb = openpyxl.load_workbook(hedef_yol)
    tablolar = tablolari_bul(wb)
    if not tablolar:
        return 0, 0
    toplam_satir, toplam_formul = 0, 0

    for (sn, tn), tbl in tablolar.items():
        ws = wb[sn]
        mn, mr, mx, sr = range_boundaries(tbl.ref)
        profil, dolu = kolon_profili(ws, mn, mr, sr, mx)
        if not profil:
            continue

        ilk_veri = mr + 1
        son_hedef = ilk_veri if tek_satir else min(ilk_veri + hedef_satir - 1, 1048575)

        # mevcut veriyi temizle (formül olmayan)
        for r in range(ilk_veri, sr + 1):
            for c in range(mn, mx + 1):
                h = ws.cell(r, c)
                if isinstance(h, MergedCell):
                    continue
                if not F(h.value):
                    h.value = None

        for c, prof in profil.items():
            for idx, r in enumerate(range(ilk_veri, son_hedef + 1)):
                h = ws.cell(r, c)
                if isinstance(h, MergedCell):
                    continue
                if prof["tip"] == "formul" and prof["formul"]:
                    kaynak_koord = f"{get_column_letter(c)}{prof['kaynak']}"
                    hedef_koord  = f"{get_column_letter(c)}{r}"
                    try:
                        h.value = Translator(prof["formul"], origin=kaynak_koord
                                             ).translate_formula(hedef_koord)
                        toplam_formul += 1
                    except Exception:
                        h.value = prof["formul"]
                else:
                    h.value = deger_uret(prof, idx, senaryo)

        # tablo referansını genişlet
        yeni_son = max(son_hedef, sr)
        tbl.ref = f"{get_column_letter(mn)}{mr}:{get_column_letter(mx)}{yeni_son}"
        toplam_satir += (son_hedef - ilk_veri + 1)

    wb.save(hedef_yol)
    return toplam_satir, toplam_formul


def hata_tara(yol):
    """Hesaplanmış dosyada kullanıcıya görünen hata değerlerini bul."""
    wb = openpyxl.load_workbook(yol, data_only=True)
    bulgular = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                s = metin(c.value).strip()
                if s in HATA_DEGERLERI:
                    bulgular.append(f"{ws.title}!{c.coordinate}={s}")
                    if len(bulgular) > 200:
                        return bulgular
    return bulgular


def karar_oku(yol, karar_sayfasi):
    wb = openpyxl.load_workbook(yol, data_only=True)
    if karar_sayfasi not in wb.sheetnames:
        return None
    ws = wb[karar_sayfasi]
    for row in ws.iter_rows():
        hucreler = list(row)
        for i, c in enumerate(hucreler):
            if metin(c.value).strip().upper() == "KARAR" and i + 1 < len(hucreler):
                v = hucreler[i + 1].value
                if v not in (None, ""):
                    return metin(v).strip()
    return None


# ===========================================================================
def olcek_testi(yol, spec):
    """Ö1 — N satırda bozulmadan çalışıyor mu, ne kadar sürede?"""
    hedef = spec["olcek_satir"]
    sinir = float(spec["azami_hesap_saniye"])
    with tempfile.TemporaryDirectory() as td:
        v = os.path.join(td, "olcek.xlsx")
        satir, formul = varyant_uret(yol, v, hedef, "normal")
        if satir == 0:
            kal("Ö1", "Excel Tablosu bulunamadı — ölçek testi yapılamadı")
            return
        sonuc, gecen = recalc(v, timeout=300)
        if sonuc is None:
            uyar("Ö1", "LibreOffice yok — ölçek testi çalıştırılamadı")
            return
        if "error" in sonuc:
            kal("Ö1", f"{satir} satırda yeniden hesaplama BAŞARISIZ: {sonuc['error']}")
            return
        hatalar = sonuc.get("total_errors", 0)
        if hatalar:
            kal("Ö1", f"{satir} satırda {hatalar} formül hatası: "
                      f"{list(sonuc.get('error_summary', {}))[:5]}")
        elif gecen > sinir:
            kal("Ö1", f"{satir} satır temiz ama yeniden hesap {gecen:.1f} sn "
                      f"(sınır {sinir} sn) — dosya kullanılamaz derecede yavaş")
        else:
            gec("Ö1", f"{satir} satır / {formul} formül · 0 hata · {gecen:.1f} sn")

        gorunen = hata_tara(v)
        if gorunen:
            kal("Ö1b", f"{len(gorunen)} hücrede kullanıcıya görünen hata: {gorunen[:5]}")
        else:
            gec("Ö1b", "Ölçekte kullanıcıya görünen hata değeri yok")


SENARYOLAR = ["tumu_sifir", "tumu_bos", "negatif_deger", "gelecek_tarih",
              "yinelenen_kayit", "tek_satir", "azami_deger"]

def uc_durum_testi(yol, spec, sadece=None):
    """Ö2 — yedi uç durumda hata görünmemeli, karar sessizce UYGUN olmamalı."""
    ksf = spec["karar_sayfasi"]
    beklenen_bos = metin(spec["bos_dosya_beklenen_karar"]).strip().upper()
    senaryolar = [sadece] if sadece else SENARYOLAR
    kirik, karar_sorunu, gecti = [], [], []

    with tempfile.TemporaryDirectory() as td:
        for sen in senaryolar:
            v = os.path.join(td, f"{sen}.xlsx")
            satir, _ = varyant_uret(yol, v, 40, sen, tek_satir=(sen == "tek_satir"))
            if satir == 0:
                continue
            sonuc, _ = recalc(v, timeout=180)
            if sonuc is None:
                uyar("Ö2", "LibreOffice yok — uç durum testi çalıştırılamadı")
                return
            if "error" in sonuc:
                kirik.append(f"{sen}: hesaplama başarısız")
                continue
            gorunen = hata_tara(v)
            if gorunen:
                kirik.append(f"{sen}: {len(gorunen)} hata hücresi ({gorunen[0]})")
                continue

            karar = karar_oku(v, ksf)
            if karar is not None:
                k = karar.upper()
                if sen == "tumu_bos" and k != beklenen_bos:
                    karar_sorunu.append(f"tumu_bos → karar '{karar}' (beklenen '{beklenen_bos}')")
                elif sen in ("negatif_deger", "azami_deger") and k == "UYGUN":
                    karar_sorunu.append(f"{sen} → karar 'UYGUN' (uç veride sessiz onay)")
                else:
                    gecti.append(f"{sen}→{karar}")
            else:
                gecti.append(f"{sen}→(karar hücresi yok)")

    if kirik:
        kal("Ö2", f"Uç durumda kırılma: {kirik}")
    elif karar_sorunu:
        kal("Ö2b", f"Uç durumda sessiz yanlış karar: {karar_sorunu}")
    else:
        gec("Ö2", f"{len(gecti)} uç durum temiz: {gecti}")


# ===========================================================================
def rapor_yaz(yol, dosya):
    kaldi = [s for s in SONUC if s[1] == "KALDI"]
    uyari = [s for s in SONUC if s[1] == "UYARI"]
    gecti = [s for s in SONUC if s[1] == "GECTI"]
    sat = ["# ÖLÇEK VE UÇ DURUM RAPORU (Ö1, Ö2)", "", f"Dosya: `{dosya}`", "",
           f"- GEÇTİ: {len(gecti)}", f"- UYARI: {len(uyari)}",
           f"- **KALDI: {len(kaldi)}**", "", "## Kalan kapılar", ""]
    for k, s, m in kaldi: sat.append(f"- **{k}** — {m}")
    sat += ["", "## Uyarılar", ""]
    for k, s, m in uyari: sat.append(f"- {k} — {m}")
    sat += ["", "## Geçen kapılar", ""]
    for k, s, m in gecti: sat.append(f"- {k} — {m}")
    with open(yol, "w", encoding="utf-8") as f:
        f.write("\n".join(sat))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("dosya")
    ap.add_argument("--spec", default=None)
    ap.add_argument("--satir", type=int, default=None)
    ap.add_argument("--sure", type=float, default=None)
    ap.add_argument("--senaryo", default=None, choices=SENARYOLAR)
    ap.add_argument("--rapor", default="RAPOR_OLCEK.md")
    ap.add_argument("--atla-olcek", action="store_true")
    a = ap.parse_args()

    spec = dict(VARSAYILAN)
    if a.spec:
        try:
            import yaml
            with open(a.spec, encoding="utf-8") as f:
                y = yaml.safe_load(f) or {}
            spec.update(y.get("denetim", {}))
            t = (y.get("testler") or {}).get("olcek") or {}
            if "satir" in t: spec["olcek_satir"] = t["satir"]
            if "azami_hesap_saniye" in t: spec["azami_hesap_saniye"] = t["azami_hesap_saniye"]
        except Exception as e:
            print(f"[uyari] spec okunamadi: {e}")

    if a.satir: spec["olcek_satir"] = a.satir
    if a.sure:  spec["azami_hesap_saniye"] = a.sure

    random.seed(20260809)   # tekrarlanabilir üretim

    if not a.atla_olcek:
        olcek_testi(a.dosya, spec)
    uc_durum_testi(a.dosya, spec, a.senaryo)

    rapor_yaz(a.rapor, a.dosya)

    kaldi = [s for s in SONUC if s[1] == "KALDI"]
    for k, s, m in SONUC:
        isaret = {"GECTI": "  GEÇTİ", "UYARI": "  UYARI", "KALDI": "! KALDI"}[s]
        print(f"{isaret}  {k:5s} {m}")
    print("\n" + "=" * 74)
    print(f"SONUÇ: {len(kaldi)} KALDI | rapor: {a.rapor}")
    print("SEVK EDİLEMEZ" if kaldi else "SEVK EDİLEBİLİR")
    sys.exit(1 if kaldi else 0)


if __name__ == "__main__":
    main()
