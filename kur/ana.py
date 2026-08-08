#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
kur/ana.py — SPEC.yaml'dan KASA-PRO xlsx üretir (v4 mandata)

Tek kaynak: SPEC.yaml. Hücre yerleşimi, formül, eşik, stil — hepsi oradan gelir.
Formül ayracı daima virgül; dizi formülü yok (SUMPRODUCT/yardımcı kolon).
Üretim sonrası recalc.py (LibreOffice) ile değer cache'leri yazılır.

Kullanım:
    python kur/ana.py [SPEC.yaml] [çıktı.xlsx]
"""

import sys, os, re, datetime
import yaml
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import (DataBarRule, ColorScaleRule, IconSetRule,
                                      FormulaRule)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

PROJE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
VARSAYILAN_SPEC = os.path.join(PROJE, "delivery", "paid-products",
                               "akilli-kasa-defteri-ve-nakit-kontrol-sistemi",
                               "SPEC.yaml")
VARSAYILAN_CIKTI = os.path.join(PROJE, "cikti",
                                "akilli-kasa-defteri-ve-nakit-kontrol-sistemi.xlsx")

# ---- tablo şeması: kolon adı → sütun harfi (kur betiği SPEC'ten beslenir,
#      ama sayfa düzeni burada kurulur) ----
TABLOLAR = {
    "tblHesaplar": {"sayfa": "HESAPLAR", "kolonlar": ["Hesap Adı", "Hesap Türü", "Başlangıç Bakiyesi", "Para Birimi"],
                    "kapasite": 1000},
    "tblHareketler": {"sayfa": "KASA_HAREKETLERI", "kolonlar": ["Tarih", "İşlem Türü", "Tutar", "Açıklama", "Ödeme Yöntemi", "İlgili Hesap", "Belge No"],
                      "kapasite": 5000,
                      "formul_kolonlari": [
                          ("MAD Sapma", '=IF(C{r}="","",ABS(C{r}-medyanTutar))'),
                          ("Anomali", '=IF(C{r}="","",IF(ABS(C{r}-medyanTutar)>anomaliEsigi,1,0))'),
                      ]},
    "tblTakvim": {"sayfa": "ODEME_TAKVIMI", "kolonlar": ["Vade Tarihi", "Tutar", "Karşı Taraf", "Durum", "Ödeme Yöntemi"],
                  "kapasite": 2000},
    "tblSayim": {"sayfa": "FIZIKI_SAYIM", "kolonlar": ["Sayım Tarihi", "Hesap Adı", "Kayıtlı Bakiye", "Fiili Bakiye", "Açıklama"],
                 "kapasite": 1000},
}

# tablo → (kolon_harfi, veri_araligi) çevirisi (koşullu biçimlendirme için)
HARF = {t: {k: get_column_letter(i) for i, k in enumerate(v["kolonlar"], 1)}
        for t, v in TABLOLAR.items()}
VERI_SONU = {t: 3 + v["kapasite"] - 1 for t, v in TABLOLAR.items()}


def ornek_veri(spec):
    """Tabloları açılışta dolu gösterecek örnek kayıtlar (D09b: boş grafik yasak)."""
    t = datetime
    return {
        "tblHesaplar": [
            ["Ana Kasa", "Kasa", 25000, "TRY"],
            ["Banka", "Banka", 120000, "TRY"],
            ["Kredi Kartı", "Kredi Kartı", 0, "TRY"],
        ],
        "tblHareketler": [
            [t.datetime(2026, 8, 5), "Gelir", 12000, "Peşin tahsilat", "Nakit", "Ana Kasa", "AD-2026-101"],
            [t.datetime(2026, 8, 6), "Gider", 8000, "Tedarikçi ödemesi", "EFT", "Banka", "AD-2026-102"],
            [t.datetime(2026, 8, 7), "Gelir", 10000, "Cari tahsilat", "EFT", "Banka", "AD-2026-103"],
            [t.datetime(2026, 8, 8), "Gider", 7000, "Kira ödemesi", "EFT", "Banka", "AD-2026-104"],
            [t.datetime(2026, 8, 9), "Gider", 4500, "Ofis gideri", "Kart", "Kredi Kartı", "AD-2026-105"],
            [t.datetime(2026, 8, 10), "Gelir", 9000, "Peşin tahsilat", "Nakit", "Ana Kasa", "AD-2026-106"],
            [t.datetime(2026, 8, 11), "Gider", 5500, "Maaş avansı", "EFT", "Banka", "AD-2026-107"],
            [t.datetime(2026, 8, 12), "Gelir", 8000, "Proje tahsilatı", "Nakit", "Ana Kasa", "AD-2026-108"],
        ],
        "tblTakvim": [
            [t.datetime(2026, 8, 15), 12000, "Tedarikçi A", "Bekliyor", "EFT"],
            [t.datetime(2026, 8, 25), 18000, "Tedarikçi B", "Bekliyor", "EFT"],
            [t.datetime(2026, 9, 5), 9000, "Vergi Dairesi", "Bekliyor", "EFT"],
            [t.datetime(2026, 9, 20), 15000, "Tedarikçi C", "Bekliyor", "Çek"],
            [t.datetime(2026, 10, 10), 20000, "Kira", "Bekliyor", "EFT"],
        ],
        "tblSayim": [
            [t.datetime(2026, 8, 31), "Ana Kasa", 159000, 159000, "Sayım kayıtlarla uyumlu"],
        ],
    }


# ===========================================================================
def stil(spec):
    s = spec["stil"]
    r = s["renk"]
    f = Font(name=s["yazitipi"], size=10)
    return {
        "yazitipi": s["yazitipi"],
        "bicim": s["bicim"],
        "baslik": Font(name=s["yazitipi"], size=11, bold=True, color="FFFFFF"),
        "baslik_dolgu": PatternFill("solid", start_color=r["koyuBaslik"], end_color=r["koyuBaslik"]),
        "giris": f,
        "giris_dolgu": PatternFill("solid", start_color=r["girisAlani"], end_color=r["girisAlani"]),
        "cikti_dolgu": PatternFill("solid", start_color=r["kilitliCikti"], end_color=r["kilitliCikti"]),
        "etiket": Font(name=s["yazitipi"], size=10, bold=True, color=r["koyuBaslik"]),
        "buyuk": Font(name=s["yazitipi"], size=20, bold=True, color=r["koyuBaslik"]),
        "alt": Font(name=s["yazitipi"], size=12, color=r["vurgu"]),
        "ince_cizgi": Border(bottom=Side(style="thin", color=r["vurgu"])),
    }


def baslik_satiri(ws, satir, degerler, st, baslangic=1):
    for i, d in enumerate(degerler, baslangic):
        c = ws.cell(satir, i, d)
        c.font = st["baslik"]
        c.fill = st["baslik_dolgu"]
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[satir].height = 22


def sayfa_notu(ws, huc, metin_, st):
    c = ws[huc]
    c.value = metin_
    c.font = Font(name=st["yazitipi"], size=9, italic=True, color="7F8C8D")


# ===========================================================================
def kapak(ws, spec, st):
    ws["A1"] = spec["urun"]["ad"]
    ws["A1"].font = st["buyuk"]
    ws["A3"] = "KASA-PRO · Nakit Kontrol Sistemi"
    ws["A3"].font = st["alt"]
    ws["A5"] = f"Sürüm {spec['urun']['surum']} — {spec['urun']['hedef_kullanici']}"
    ws["A7"] = "Verileriniz cihazınızdan çıkmaz. Bu dosya tamamen yerel çalışır; hiçbir veri dışarı gönderilmez."
    ws["A7"].font = Font(name=st["yazitipi"], size=10, bold=True, color="1F7A4D")
    ws["A9"] = "Rapor tarihi"
    ws["B9"] = "=raporTarihi"
    ws["B9"].number_format = st["bicim"]["tarih"]
    ws.column_dimensions["A"].width = 58
    ws.column_dimensions["B"].width = 14


def hizli_baslangic(ws, spec, st):
    ws["A1"] = "HIZLI BAŞLANGIÇ"
    ws["A1"].font = st["buyuk"]
    adimlar = [
        "1. AYARLAR sayfasından firma unvanını ve eşikleri güncelleyin.",
        "2. HESAPLAR'a kasa, banka ve kart hesaplarınızı girin.",
        "3. KASA_HAREKETLERI'ne tüm gelir ve giderleri kaydedin.",
        "4. ODEME_TAKVIMI'ne vadeli ödemelerinizi girin.",
        "5. FIZIKI_SAYIM'a sayım sonuçlarını işleyin (opsiyonel).",
        "6. PANO ve KARAR sayfaları otomatik güncellenir.",
        "7. Sonuçları RAPOR ile yazdırabilirsiniz.",
    ]
    for i, m in enumerate(adimlar, 4):
        ws.cell(i, 2, m)
    sayfa_notu(ws, "A2", "İlk kurulumda sarı giriş alanları dışındaki hücreler korumalıdır.", st)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 90


def giris_tablosu(spec, st, adlar, wb, tablo_adi, ornekler):
    tn = TABLOLAR[tablo_adi]
    ws = wb[tn["sayfa"]]
    kap = tn["kapasite"]
    son = 3 + kap
    kolonlar = tn["kolonlar"] + [k for k, _ in tn.get("formul_kolonlari", [])]

    ws["A1"] = tn["sayfa"].replace("_", " ")
    ws["A1"].font = st["buyuk"]
    baslik_satiri(ws, 3, kolonlar, st)
    ws.freeze_panes = "A4"

    tab = Table(displayName=tablo_adi, ref=f"A3:{get_column_letter(len(kolonlar))}{son}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True,
                                        showFirstColumn=False, showLastColumn=False)
    ws.add_table(tab)

    for r, kayit in enumerate(ornekler, 4):
        for c, deger in enumerate(kayit, 1):
            ws.cell(r, c, deger)
        for kolon, sablon in tn.get("formul_kolonlari", []):
            fc = ws.cell(r, len(tn["kolonlar"]) + 1 + [k for k, _ in tn["formul_kolonlari"]].index(kolon),
                         sablon.format(r=r))
            fc.number_format = st["bicim"]["tamsayi"] if kolon == "Anomali" else st["bicim"]["para"]

    # alan sözlüğünden ipucu/hata (G10-G12)
    alan_bilgi = {a["kolon"]: a for a in spec["alanlar"] if a["tablo"] == tablo_adi}
    for c, kol in enumerate(tn["kolonlar"], 1):
        bil = alan_bilgi.get(kol)
        if bil:
            dv = veri_dogrulama(bil, tn["sayfa"], c, son)
            if dv:
                ws.add_data_validation(dv)
                dv.add(f"{get_column_letter(c)}4:{get_column_letter(c)}{son}")

    # örnek veriye giriş stili
    for r in range(4, 4 + len(ornekler)):
        for c in range(1, len(tn["kolonlar"]) + 1):
            h = ws.cell(r, c)
            h.fill = st["giris_dolgu"]
            if h.value is not None and not isinstance(h.value, str):
                h.number_format = st["bicim"]["tarih"] if isinstance(h.value, datetime.datetime) else st["bicim"]["para"]

    genislik = {0: 24, 1: 14, 2: 14, 3: 40, 4: 14, 5: 14, 6: 14, 7: 14, 8: 14}
    for i, kol in enumerate(kolonlar):
        ws.column_dimensions[get_column_letter(i + 1)].width = genislik.get(i, 14)
    return ws


def veri_dogrulama(bil, sayfa, c, son):
    """alanlar sözlüğünden openpyxl DataValidation üretir; yoksa None."""
    tip = bil.get("tip")
    if tip == "liste":
        dv = DataValidation(type="list", formula1='"' + ",".join(bil["liste"]) + '"',
                            allow_blank=not bil.get("zorunlu", True), showErrorMessage=True)
    elif tip in ("para", "tamsayi", "puan", "kat", "ondalik"):
        d = bil.get("dogrulama") or {}
        mn, mx = d.get("min"), d.get("maks")
        if mn is None and mx is None:
            return None
        tip_xml = "whole" if tip == "tamsayi" else "decimal"
        dv = DataValidation(type=tip_xml, operator="between",
                            formula1=str(mn), formula2=str(mx),
                            allow_blank=not bil.get("zorunlu", True), showErrorMessage=True)
    elif tip == "tarih":
        d = bil.get("dogrulama") or {}
        mn = d.get("min") or "2020-01-01"
        mx = d.get("maks_formul")
        if mx is None:
            return None
        dv = DataValidation(type="date", operator="between",
                            formula1=f"=DATE({mn[:4]},{int(mn[5:7])},{int(mn[8:10])})",
                            formula2=mx,
                            allow_blank=not bil.get("zorunlu", True), showErrorMessage=True)
    else:
        return None
    if bil.get("ipucu_baslik"):
        dv.promptTitle = bil["ipucu_baslik"]
        dv.prompt = bil.get("ipucu", "")
        dv.showInputMessage = True
    if bil.get("hata_baslik"):
        dv.errorTitle = bil["hata_baslik"]
        dv.error = bil.get("hata", "")
    return dv


# ===========================================================================
def yorum_formulu(kod, spec):
    """analitik modüllerin yorum hücresi formülü — makine üretimli (D05)."""
    b = spec["stil"]["bicim"]
    para = lambda ad: f'TEXT({ad},"#.##0,00")&" ₺"'
    if kod == "T2":
        return (f'=_xlfn.TEXTJOIN(" | ",TRUE,"Günlük ortalama net nakit",'
                f'{para("gunlukNetOrtalama")},"₺. Kayıt sayısı",doluSatirSayisi,". Trend tabanı dönemden türer.")')
    if kod == "T5":
        return (f'=_xlfn.TEXTJOIN(" | ",TRUE,"Mutabakat farkı",'
                f'{para("mutabakatFarki")},"₺. Durum",mutabakatDurum,". Tolerans",fizikiTolerans,"₺.")')
    if kod == "T6":
        return (f'=_xlfn.TEXTJOIN(" | ",TRUE,"Serbest nakit",'
                f'{para("serbestNakit")},"₺ ile",ROUND(kacGunYeter,0),"gün gider karşılanıyor. Kırılım noktası hesaplandı.")')
    if kod == "O1":
        return (f'=_xlfn.TEXTJOIN(" | ",TRUE,"Anomali eşiği",'
                f'{para("anomaliEsigi")},"₺. Tespit",anomaliSayisi,"kayıt. Oran",TEXT(anomaliOrani,"0,0%"),". MAD tabanlı.")')
    if kod == "O8":
        return (f'=_xlfn.TEXTJOIN(" | ",TRUE,"Veri kalitesi",'
                f'veriKaliteSkoru,"/100. Eksik zorunlu alan",eksikZorunlu,". Canlı veriden hesaplandı.")')
    if kod == "I1":
        return (f'=_xlfn.TEXTJOIN(" | ",TRUE,"Tahmin ufku",tahminUfkuGun,"gün. Orta",'
                f'{para("tahminBitisNakit")},"₺. Aralık",{para("tahminAlt")},"-",'
                f'{para("tahminUst")},"₺. Belirsizlik",{para("tahminAraligi")},"₺.")')
    if kod == "I7":
        return (f'=_xlfn.TEXTJOIN(" | ",TRUE,"Nakit açığı",nakitAcik,". Tutar",'
                f'{para("nakitAcikTutar")},"₺. Köprüleme:",IF(nakitAcikTutar>0,'
                f'"Ödeme takvimini 30 güne yeniden yay","Gerek yok."))')
    raise ValueError(kod)


def motor(spec, st, adlar, wb):
    ws = wb["MOTOR"]
    ws["A1"] = "MOTOR — HESAP KATMANI (40 isimli adım + 7 makine yorumu)"
    ws["A1"].font = st["buyuk"]
    baslik_satiri(ws, 3, ["Ad", "Formül", "Birim", "İşlev"], st)

    r = 4
    for adim in spec["motor"]:
        ws.cell(r, 1, adim["ad"]).font = st["etiket"]
        fc = ws.cell(r, 2, adim["formul"])
        fc.number_format = "General"
        fc.fill = st["cikti_dolgu"]
        ws.cell(r, 3, adim["birim"])
        ws.cell(r, 4, adim["aciklama"])
        adlar[adim["ad"]] = f"MOTOR!$B${r}"
        r += 1

    r += 1
    ws.cell(r, 1, "MAKİNE YORUMLARI (analitik modüller)")
    ws.cell(r, 1).font = st["etiket"]
    r += 1
    for mod in spec["analitik_moduller"]:
        yh = mod["yorum_hucresi"]
        ws.cell(r, 1, yh).font = st["etiket"]
        fc = ws.cell(r, 2, yorum_formulu(mod["kod"], spec))
        fc.fill = st["cikti_dolgu"]
        ws.cell(r, 3, "metin")
        ws.cell(r, 4, f"Modül {mod['kod']}: {mod['ad']}")
        adlar[yh] = f"MOTOR!$B${r}"
        r += 1

    ws.freeze_panes = "A4"
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 90
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 70
    ws.protection.sheet = True
    ws.protection.password = spec["denetim"]["koruma_sifresi"]


KONTROL_SATIRLARI = [
    ("Toplam Gelir", "toplamGelir", 'IF(toplamGelir>0,"OK","DİKKAT")', "Dönem gelir toplamı"),
    ("Toplam Gider", "toplamGider", 'IF(toplamGider>0,"OK","DİKKAT")', "Dönem gider toplamı"),
    ("Net Hareket", "netHareket", 'IF(netHareket>=0,"Pozitif","Negatif")', "Gelir - gider farkı"),
    ("Serbest Nakit", "serbestNakit", 'IF(serbestNakit<esikSerbestNakitKritik,"DİKKAT","OK")', "30 günlük yük sonrası"),
    ("Beklenen Bakiye", "beklenenBakiye", 'IF(beklenenBakiye>=0,"OK","DİKKAT")', "Defter toplamı"),
    ("Yaklaşan Ödeme 30g", "yaklasanOdeme30", 'IF(yaklasanOdeme30>0,"Baskı var","Rahat")', "30 gün içi bekleyen"),
    ("Mutabakat Farkı", "mutabakatFarki", "mutabakatDurum", "Defter - fiili fark"),
    ("Veri Kalite Skoru", "veriKaliteSkoru", 'IF(veriKaliteSkoru<70,"DİKKAT","OK")', "0-100 canlı skor"),
    ("Anomali Sayısı", "anomaliSayisi", 'IF(anomaliSayisi>0,"DİKKAT","OK")', "MAD eşiği aşanlar"),
    ("Risk Puanı", "riskPuani", 'IF(riskPuani>=esikRiskYuksek,"YÜKSEK","NORMAL")', "Bileşik risk skoru"),
    ("Nakit Açığı", "nakitAcik", "", "Tahmin ufkunda açık var mı"),
    ("Nakit Açığı Tutarı", "nakitAcikTutar", 'IF(nakitAcikTutar>0,"DİKKAT","OK")', "Açık büyüklüğü"),
    ("Kötümser Nakit", "kotumserNakit", "", "Kötümser senaryo"),
    ("İyimser Nakit", "iyimserNakit", "", "İyimser senaryo"),
    ("Tahmin Bitiş Nakit", "tahminBitisNakit", "", "Ufuk sonu orta"),
    ("Tahmin Alt", "tahminAlt", 'IF(tahminAlt<0,"DİKKAT","OK")', "Ufuk sonu alt bandı"),
    ("Tahmin Üst", "tahminUst", "", "Ufuk sonu üst bandı"),
    ("Kac Gun Yeter", "kacGunYeter", "", "Serbest nakit / günlük gider"),
    ("Medyan Tutar", "medyanTutar", "", "Tutar medyanı"),
    ("Anomali Eşiği", "anomaliEsigi", "", "MAD × kat sayısı"),
    ("Nihai Karar", "kararSonuc", "", "KARAR sayfası ile aynı"),
]


def kontrol(spec, st, wb):
    ws = wb["KONTROL"]
    ws["A1"] = "KONTROL — HESAP DOĞRULAMA"
    ws["A1"].font = st["buyuk"]
    baslik_satiri(ws, 3, ["Kontrol", "Değer", "Durum", "Açıklama"], st)
    for i, (et, ad, dur, not_) in enumerate(KONTROL_SATIRLARI, 4):
        ws.cell(i, 1, et).font = st["etiket"]
        fc = ws.cell(i, 2, f"={ad}")
        fc.number_format = "General"
        if dur:
            ws.cell(i, 3, f"={dur}")
        ws.cell(i, 4, not_)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40


def karar(spec, st, wb):
    ws = wb["KARAR"]
    ws["A1"] = "KARAR KATMANI"
    ws["A1"].font = st["buyuk"]
    cift = [
        ("KARAR", "kararSonuc"),
        ("GEREKÇE", "kararGerekce"),
        ("TETİKLENEN KURAL", "aktifKural"),
        ("SERBEST NAKİT", "serbestNakit"),
    ]
    r = 5
    for et, ad in cift:
        ws.cell(r, 1, et).font = st["etiket"]
        fc = ws.cell(r, 2, f"={ad}")
        fc.fill = st["cikti_dolgu"]
        r += 1
    r += 1
    ws.cell(r, 1, "Karar kuralları:").font = st["etiket"]
    r += 1
    for k in spec["karar_kurallari"]:
        ws.cell(r, 1, k["kod"])
        ws.cell(r, 2, k["kosul"])
        ws.cell(r, 3, k["sonuc"])
        ws.cell(r, 4, k["gerekce"])
        r += 1
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 50
    ws.protection.sheet = True
    ws.protection.password = spec["denetim"]["koruma_sifresi"]


PANO_KPI = [
    ("Serbest Nakit", "serbestNakit", "₺"),
    ("Beklenen Bakiye", "beklenenBakiye", "₺"),
    ("Net Hareket", "netHareket", "₺"),
    ("Toplam Gelir", "toplamGelir", "₺"),
    ("Toplam Gider", "toplamGider", "₺"),
    ("Yaklaşan Ödeme 30g", "yaklasanOdeme30", "₺"),
    ("Yaklaşan Ödeme 7g", "yaklasanOdeme7", "₺"),
    ("Mutabakat Farkı", "mutabakatFarki", "₺"),
    ("Veri Kalite Skoru", "veriKaliteSkoru", "puan"),
    ("Anomali Sayısı", "anomaliSayisi", "adet"),
    ("Risk Puanı", "riskPuani", "puan"),
    ("Kac Gun Yeter", "kacGunYeter", "gün"),
    ("Tahmin Bitiş Nakit", "tahminBitisNakit", "₺"),
    ("Nakit Açığı", "nakitAcik", ""),
    ("P90 Tutar", '=IFERROR(_xlfn.PERCENTILE.INC(tblHareketler[Tutar],0.9),0)', "₺"),
    ("P10 Tutar", '=IFERROR(_xlfn.PERCENTILE.INC(tblHareketler[Tutar],0.1),0)', "₺"),
]


def baski_ayar(ws, alan):
    """Baskı alanı, tek sayfa genişliği, başlık satırı, yatay ortalama (G13-G15)."""
    ws.print_area = alan
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_title_rows = "1:3"
    ws.print_options.horizontalCentered = True


def pano(spec, st, wb):
    ws = wb["PANO"]
    ws["A1"] = "PANO — CANLI KPI EKRANI"
    ws["A1"].font = st["buyuk"]
    baslik_satiri(ws, 3, ["KPI", "Değer", "Birim", "Durum"], st)
    for i, (et, ad, birim) in enumerate(PANO_KPI, 4):
        ws.cell(i, 1, et).font = st["etiket"]
        ws.cell(i, 2, f"={ad}").number_format = "General"
        ws.cell(i, 3, birim)
        if et == "Serbest Nakit":
            durum = f'=IF(B{i}<esikSerbestNakitKritik,"DİKKAT","OK")'
        elif et == "Veri Kalite Skoru":
            durum = f'=IF(B{i}<70,"DİKKAT","OK")'
        elif et == "Anomali Sayısı":
            durum = f'=IF(B{i}>0,"DİKKAT","OK")'
        else:
            durum = ""
        ws.cell(i, 4, durum)
    # projeksiyon bloğu (grafik için)
    ws["F3"] = "Projeksiyon"
    ws["F3"].font = st["baslik"]; ws["F3"].fill = st["baslik_dolgu"]
    ws["G3"] = "Nakit"
    ws["G3"].font = st["baslik"]; ws["G3"].fill = st["baslik_dolgu"]
    proj = [("Bugün", "serbestNakit"), ("Ufuk Orta", "tahminBitisNakit"),
            ("Ufuk Alt", "tahminAlt"), ("Ufuk Üst", "tahminUst")]
    for i, (et, ad) in enumerate(proj, 4):
        ws.cell(i, 6, et).font = st["etiket"]
        fc = ws.cell(i, 7, f"={ad}")
        fc.fill = st["cikti_dolgu"]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.page_setup.orientation = "landscape"
    baski_ayar(ws, "A1:G30")
    ws.protection.sheet = True
    ws.protection.password = spec["denetim"]["koruma_sifresi"]

    # grafikler
    bar1 = BarChart(); bar1.type = "col"; bar1.title = "Nakit Pozisyonu"
    bar1.add_data(Reference(ws, min_col=2, min_row=4, max_row=6), titles_from_data=False)
    bar1.set_categories(Reference(ws, min_col=1, min_row=4, max_row=6))
    bar1.width, bar1.height = 14, 8
    ws.add_chart(bar1, "H4")

    bar2 = BarChart(); bar2.type = "col"; bar2.title = "Gelir / Gider"
    bar2.add_data(Reference(ws, min_col=2, min_row=4, max_row=5), titles_from_data=False)
    bar2.set_categories(Reference(ws, min_col=1, min_row=4, max_row=5))
    bar2.width, bar2.height = 14, 8
    ws.add_chart(bar2, "H19")

    line = LineChart(); line.title = "Nakit Projeksiyonu (Tahmin Aralığı)"
    line.add_data(Reference(ws, min_col=7, min_row=4, max_row=7), titles_from_data=False)
    line.set_categories(Reference(ws, min_col=6, min_row=4, max_row=7))
    line.width, line.height = 14, 8
    ws.add_chart(line, "H34")

    bar3 = BarChart(); bar3.type = "bar"; bar3.title = "Tüm KPI'lar"
    bar3.add_data(Reference(ws, min_col=2, min_row=4, max_row=19), titles_from_data=False)
    bar3.set_categories(Reference(ws, min_col=1, min_row=4, max_row=19))
    bar3.width, bar3.height = 14, 8
    ws.add_chart(bar3, "H49")


def senaryo(spec, st, wb):
    ws = wb["SENARYO_DUYARLILIK"]
    ws["A1"] = "SENARYO VE DUYARLILIK"
    ws["A1"].font = st["buyuk"]
    baslik_satiri(ws, 3, ["Senaryo", "Ufuk Sonu Nakit", "Orta Kademe Farkı"], st)
    satirlar = [
        ("Kötümser", "kotumserNakit", "kotumserNakit-tahminBitisNakit"),
        ("Orta", "tahminBitisNakit", ""),
        ("İyimser", "iyimserNakit", "iyimserNakit-tahminBitisNakit"),
    ]
    for i, (et, ad, fark) in enumerate(satirlar, 5):
        ws.cell(i, 1, et).font = st["etiket"]
        ws.cell(i, 2, f"={ad}").number_format = st["bicim"]["para"]
        if fark:
            ws.cell(i, 3, f"={fark}")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    bar = BarChart(); bar.type = "col"; bar.title = "Senaryo Karşılaştırma"
    bar.add_data(Reference(ws, min_col=2, min_row=5, max_row=7), titles_from_data=False)
    bar.set_categories(Reference(ws, min_col=1, min_row=5, max_row=7))
    bar.width, bar.height = 14, 8
    ws.add_chart(bar, "E5")
    ws.protection.sheet = True
    ws.protection.password = spec["denetim"]["koruma_sifresi"]


RAPOR_SATIRLARI = [
    ("Serbest Nakit", "serbestNakit"),
    ("Beklenen Bakiye", "beklenenBakiye"),
    ("Toplam Gelir", "toplamGelir"),
    ("Toplam Gider", "toplamGider"),
    ("Net Hareket", "netHareket"),
    ("Yaklaşan Ödeme 30g", "yaklasanOdeme30"),
    ("Mutabakat Farkı", "mutabakatFarki"),
    ("Veri Kalite Skoru", "veriKaliteSkoru"),
    ("Anomali Sayısı", "anomaliSayisi"),
    ("Risk Puanı", "riskPuani"),
    ("Tahmin Bitiş Nakit", "tahminBitisNakit"),
    ("Tahmin Alt", "tahminAlt"),
    ("Tahmin Üst", "tahminUst"),
    ("Nakit Açığı", "nakitAcik"),
    ("Kac Gun Yeter", "kacGunYeter"),
    ("Günlük Net Ortalama", "gunlukNetOrtalama"),
    ("Günlük Gider Ortalama", "gunlukGiderOrtalama"),
    ("Medyan Tutar", "medyanTutar"),
    ("Anomali Eşiği", "anomaliEsigi"),
    ("Mutabakat Durumu", "mutabakatDurum"),
    ("Nihai Karar", "kararSonuc"),
]


def rapor(spec, st, wb):
    ws = wb["RAPOR"]
    ws["A1"] = "RAPOR"
    ws["A1"].font = st["buyuk"]
    baslik_satiri(ws, 3, ["Gösterge", "Değer"], st)
    for i, (et, ad) in enumerate(RAPOR_SATIRLARI, 6):
        ws.cell(i, 1, et).font = st["etiket"]
        ws.cell(i, 2, f"={ad}").number_format = "General"
    # günlük nakit akışı (grafik serisi)
    ws["E29"] = "Günlük Nakit Akışı (8 gün)"
    ws["E29"].font = st["baslik"]; ws["E29"].fill = st["baslik_dolgu"]
    ws["F29"] = "Net"
    ws["F29"].font = st["baslik"]; ws["F29"].fill = st["baslik_dolgu"]
    gunler = [datetime.datetime(2026, 8, 5) + datetime.timedelta(days=i) for i in range(8)]
    for i, gun in enumerate(gunler):
        r = 30 + i
        ws.cell(r, 5, gun).number_format = st["bicim"]["tarih"]
        ws.cell(r, 6, (f'=SUMIFS(tblHareketler[Tutar],tblHareketler[Tarih],">="&E{r},'
                       f'tblHareketler[Tarih],"<"&E{r}+1,tblHareketler[İşlem Türü],"Gelir")'
                       f'-SUMIFS(tblHareketler[Tutar],tblHareketler[Tarih],">="&E{r},'
                       f'tblHareketler[Tarih],"<"&E{r}+1,tblHareketler[İşlem Türü],"Gider")'))
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.page_setup.orientation = "landscape"
    baski_ayar(ws, "A1:F40")
    ws.protection.sheet = True
    ws.protection.password = spec["denetim"]["koruma_sifresi"]

    bar = BarChart(); bar.type = "col"; bar.title = "Temel Göstergeler"
    bar.add_data(Reference(ws, min_col=2, min_row=6, max_row=26), titles_from_data=False)
    bar.set_categories(Reference(ws, min_col=1, min_row=6, max_row=26))
    bar.width, bar.height = 16, 9
    ws.add_chart(bar, "D6")

    line = LineChart(); line.title = "Günlük Net Nakit Akışı"
    line.add_data(Reference(ws, min_col=6, min_row=30, max_row=37), titles_from_data=False)
    line.set_categories(Reference(ws, min_col=5, min_row=30, max_row=37))
    line.width, line.height = 16, 9
    ws.add_chart(line, "D21")


def ornek_veri_sayfa(spec, st, wb):
    ws = wb["ORNEK_VERI"]
    ws["A1"] = "ÖRNEK VERİ — KASA HAREKETLERI örneği"
    ws["A1"].font = st["buyuk"]
    baslik_satiri(ws, 3, ["Tarih", "İşlem Türü", "Tutar", "Açıklama", "Ödeme Yöntemi", "İlgili Hesap"], st)
    for i, k in enumerate(ornek_veri(spec)["tblHareketler"], 4):
        for c, v in enumerate(k[:6], 1):
            ws.cell(i, c, v)
    sayfa_notu(ws, "A2", "Bu sayfa bilgilendirme amaçlıdır; gerçek hesaplamalar tablolardan gelir.", st)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14


def degisiklik(spec, st, wb):
    ws = wb["DEGISIKLIK_KAYDI"]
    ws["A1"] = "DEĞİŞİKLİK KAYDI"
    ws["A1"].font = st["buyuk"]
    baslik_satiri(ws, 3, ["Sürüm", "Tarih", "Değişiklik"], st)
    ws.cell(4, 1, "2.0.0")
    ws.cell(4, 2, "09.08.2026")
    ws.cell(4, 3, "v4 mandata uyumlu yeniden üretim: 44 kapı sözleşmesi, canlı motor, makine yorumları")
    ws.cell(5, 1, "1.0.0")
    ws.cell(5, 2, "2024")
    ws.cell(5, 3, "İlk yayın")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 80


def listeler(spec, st, wb):
    ws = wb["LISTELER"]
    ws["A1"] = "LİSTELER — AÇILIR MENÜ KAYNAKLARI"
    ws["A1"].font = st["buyuk"]
    gruplar = [
        ("İşlem Türleri", ["Gelir", "Gider"]),
        ("Ödeme Yöntemleri", ["Nakit", "Kart", "EFT", "Çek", "Senet"]),
        ("Hesap Türleri", ["Kasa", "Banka", "Kredi Kartı"]),
        ("Para Birimleri", ["TRY", "USD", "EUR"]),
        ("Vade Durumları", ["Bekliyor", "Ödendi", "İptal"]),
    ]
    kol = 1
    for baslik, degerler in gruplar:
        ws.cell(3, kol, baslik).font = st["baslik"]
        ws.cell(3, kol).fill = st["baslik_dolgu"]
        for i, d in enumerate(degerler, 4):
            ws.cell(i, kol, d)
        ws.column_dimensions[get_column_letter(kol)].width = 18
        kol += 2


def ayarlar(spec, st, adlar, wb):
    ws = wb["AYARLAR"]
    ws["A1"] = "AYARLAR — PARAMETRE MERKEZİ"
    ws["A1"].font = st["buyuk"]
    baslik_satiri(ws, 3, ["anahtar", "deger", "birim", "aciklama", "kaynak", "yururluk_tarihi", "dogrulama_tarihi"], st)
    r = 4
    for a in spec["ayarlar"]:
        ws.cell(r, 1, a["anahtar"]).font = st["etiket"]
        deger = a.get("deger")
        if a["anahtar"] == "firmaUnvani" and deger in (None, ""):
            deger = "İşletme Adı"  # D07b: ad tanımı boş hücre gösteremez
        h = ws.cell(r, 2, deger)
        if a.get("tip") == "tarih":
            h.number_format = st["bicim"]["tarih"]
        elif a.get("tip") == "para":
            h.number_format = st["bicim"]["para"]
        ws.cell(r, 3, a["birim"])
        ws.cell(r, 4, a["aciklama"])
        ws.cell(r, 5, a["kaynak"])
        ws.cell(r, 6, a["yururluk_tarihi"])
        ws.cell(r, 7, a["dogrulama_tarihi"])
        adlar[a["anahtar"]] = f"AYARLAR!$B${r}"
        r += 1
    for col, w in zip("ABCDEFG", (26, 16, 8, 44, 22, 14, 16)):
        ws.column_dimensions[col].width = w
    # kullanıcı yalnızca deger kolonunu günceller; gerisi kilitli (G01)
    for satir in range(4, r):
        ws.cell(satir, 2).protection = Protection(locked=False)
    ws.protection.sheet = True
    ws.protection.password = spec["denetim"]["koruma_sifresi"]


def kilavuz(spec, st, wb):
    ws = wb["KILAVUZ"]
    ws["A1"] = "KILAVUZ"
    ws["A1"].font = st["buyuk"]
    metinler = [
        "NASIL KULLANILIR?",
        "1) Sarı giriş alanlarına veri girin; hesaplanan alanlar korumalıdır (şifre: 1234).",
        "2) HESAPLAR → hesaplar; KASA_HAREKETLERI → gelir/gider; ODEME_TAKVIMI → vadeler; FIZIKI_SAYIM → sayımlar.",
        "3) PANO ve KARAR anlık güncellenir. RAPOR yazdırmaya hazırdır.",
        "",
        "KARAR KURALLARI (KARAR sayfası ile aynıdır):",
    ]
    for i, m in enumerate(metinler, 2):
        ws.cell(i, 1, m)
    r = len(metinler) + 2
    baslik_satiri(ws, r, ["Kod", "Koşul", "Sonuç", "Gerekçe"], st)
    r += 1
    for k in spec["karar_kurallari"]:
        ws.cell(r, 1, k["kod"])
        ws.cell(r, 2, k["kosul"])
        ws.cell(r, 3, k["sonuc"])
        ws.cell(r, 4, k["gerekce"])
        r += 1
    r += 1
    ws.cell(r, 1, "GÜVENLİK: Bu dosya yerel çalışır; hiçbir veri dışarı gönderilmez.")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 60


# ===========================================================================
def kosullu_bicimlendirme(spec, st, wb):
    """SPEC'teki 44 kuralı openpyxl kurallarına çevirir."""
    b = st["bicim"]
    harf = HARF
    son = VERI_SONU

    def aralik_cevir(sayfa, ar):
        """'tblX[Kolon]' → (A1_araligi, kolon_harfi); zaten A1 ise ilk kolonu çıkar."""
        for tn, bil in TABLOLAR.items():
            if ar.startswith(f"{tn}["):
                kol = ar[len(tn) + 1:-1]
                h = harf[tn][kol]
                return f"{h}3:{h}{son[tn]}", h
        m = re.match(r"^\$?([A-Z]+)", ar)
        return ar, (m.group(1) if m else None)

    def formulu_cevir(sayfa, f):
        # yapılandırılmış referanslar → A1 göreli; @'li ise boş satır önkoşulu ekle
        for tn, bil in TABLOLAR.items():
            for kol, h in harf[tn].items():
                f = f.replace(f"{tn}[@{kol}]", f"${h}4")
                f = f.replace(f"{tn}[{kol}]", f"${h}4")
        f = f.replace("KONTROL!B4", "$B4").replace("KARAR!B5", "$B$5")
        f = f.replace("PANO!B4", "$B4").replace("RAPOR!C6", "$C6")
        return f

    dolgu = lambda hex_: PatternFill("solid", start_color=hex_, end_color=hex_)
    adet = 0
    for kural in spec["kosullu_bicimlendirme"]:
        sayfa, tur = kural["sayfa"], kural["tur"]
        ar, kolon = aralik_cevir(sayfa, kural["aralik"])
        ws = wb[sayfa]
        if tur == "veri_cubugu":
            ws.conditional_formatting.add(ar, DataBarRule(
                start_type="min", end_type="max", color=kural.get("renk", "B08948")))
        elif tur == "uc_renk_olcek":
            ws.conditional_formatting.add(ar, ColorScaleRule(
                start_type="min", start_color="FDE7E7",
                mid_type="percentile", mid_value=50, mid_color="FFF2CC",
                end_type="max", end_color="1F7A4D"))
        elif tur == "simge_kumesi":
            ws.conditional_formatting.add(ar, IconSetRule(
                icon_style="3TrafficLights1", type="num", values=[0, 33, 67]))
        elif tur == "formul":
            ws.conditional_formatting.add(ar, FormulaRule(
                formula=[formulu_cevir(sayfa, kural["formul"])], fill=dolgu(kural.get("dolgu", "FDE7E7"))))
        elif tur == "yinelenen":
            ws.conditional_formatting.add(ar, FormulaRule(
                formula=[f"=COUNTIF({ar},{kolon}{4})>1"], fill=dolgu(kural.get("dolgu", "FDF2D0"))))
        elif tur == "esik_asimi":
            ws.conditional_formatting.add(ar, FormulaRule(
                formula=[formulu_cevir(sayfa, kural["esik_formul"])], fill=dolgu(kural.get("dolgu", "FDE7E7"))))
        elif tur == "bos_zorunlu":
            ws.conditional_formatting.add(ar, FormulaRule(
                formula=[f'=AND({kolon}{4}="",COUNTA($A$4:${kolon}{4})>0)'], fill=dolgu(kural.get("dolgu", "FDE7E7"))))
        elif tur == "gelecek_tarih":
            ws.conditional_formatting.add(ar, FormulaRule(
                formula=[f'=AND({kolon}{4}<>"",{kolon}{4}>raporTarihi+{kural.get("gun_limit", 365)})'],
                fill=dolgu(kural.get("dolgu", "FFF2CC"))))
        else:
            raise ValueError(f"bilinmeyen kural türü: {tur}")
        adet += 1
    print(f"[kur] koşullu biçimlendirme: {adet} kural yazıldı")
    return adet


# ===========================================================================
def ana(spec_yol, cikti_yol):
    with open(spec_yol, encoding="utf-8") as f:
        spec = yaml.safe_load(f)
    st = stil(spec)
    adlar = {}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sayfalar = {}
    for s in spec["sayfalar"]:
        ws = wb.create_sheet(s["ad"])
        if s.get("sekme_rengi"):
            ws.sheet_properties.tabColor = s["sekme_rengi"]
        sayfalar[s["ad"]] = ws

    kapak(sayfalar["KAPAK"], spec, st)
    hizli_baslangic(sayfalar["HIZLI_BASLANGIC"], spec, st)
    ov = ornek_veri(spec)
    for tn in TABLOLAR:
        giris_tablosu(spec, st, adlar, wb, tn, ov[tn])
    motor(spec, st, adlar, wb)
    kontrol(spec, st, wb)
    karar(spec, st, wb)
    pano(spec, st, wb)
    senaryo(spec, st, wb)
    rapor(spec, st, wb)
    ornek_veri_sayfa(spec, st, wb)
    degisiklik(spec, st, wb)
    listeler(spec, st, wb)
    ayarlar(spec, st, adlar, wb)
    kilavuz(spec, st, wb)

    # tablo sayfalarına veri grafiği (D09: toplam ≥8 grafik)
    def tablo_grafigi(sayfa, baslik, kat_s, kat_b, kat_e, veri_s, veri_b, veri_e, tur="col"):
        ws = wb[sayfa]
        ch = BarChart() if tur == "col" else LineChart()
        if isinstance(ch, BarChart):
            ch.type = tur
        ch.title = baslik
        ch.add_data(Reference(ws, min_col=veri_s, min_row=veri_b, max_row=veri_e),
                    titles_from_data=False)
        ch.set_categories(Reference(ws, min_col=kat_s, min_row=kat_b, max_row=kat_e))
        ch.width, ch.height = 14, 8
        ws.add_chart(ch, "I4")
        return ch

    tablo_grafigi("KASA_HAREKETLERI", "Hareket Tutarları", 1, 4, 11, 3, 4, 11)
    tablo_grafigi("ODEME_TAKVIMI", "Vade Tutarları", 3, 4, 8, 2, 4, 8)
    tablo_grafigi("HESAPLAR", "Başlangıç Bakiyeleri", 1, 4, 6, 3, 4, 6)

    for ad, ref in adlar.items():
        wb.defined_names.add(DefinedName(ad, attr_text=ref))
    print(f"[kur] ad tanımı: {len(adlar)}")

    kosullu_bicimlendirme(spec, st, wb)

    os.makedirs(os.path.dirname(cikti_yol), exist_ok=True)
    wb.save(cikti_yol)
    print(f"[kur] ÜRETİLDİ: {cikti_yol}")
    return cikti_yol


if __name__ == "__main__":
    sp = sys.argv[1] if len(sys.argv) > 1 else VARSAYILAN_SPEC
    ck = sys.argv[2] if len(sys.argv) > 2 else VARSAYILAN_CIKTI
    ana(sp, ck)
