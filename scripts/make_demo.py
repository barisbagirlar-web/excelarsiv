#!/usr/bin/env python3
"""Excel Arşiv — akilli-kasa-defteri demo üretimi (openpyxl, repo dışı araç).

Girdi sheet: açık hücreler (sarı dolgu = kullanıcı girer).
Hesap / Çıktı sheet: örnek formüller (SUM, IF) + kilitli sheet koruması.
Makro içermez (.xlsx, .xlsm değil) — SmartScreen uyarısı üretmez.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Protection

OUT = (
    Path(__file__).resolve().parents[1]
    / "public"
    / "demo"
    / "akilli-kasa-defteri-demo-v3.2.xlsx"
)

INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")  # sarı dolgu = kullanıcı girer
LABEL_FONT = Font(bold=True)

wb = Workbook()

# Girdi — açık hücreler (locked=False)
ws = wb.active
ws.title = "Girdi"
girdi_rows = [
    ("Kasa Adı", "Ana Kasa"),
    ("Dönem Başı Bakiye", 0),
    ("Dönem Geliri", 0),
    ("Dönem Gideri", 0),
    ("Açıklama", "Demo — ücretli sürümde temizlenir"),
]
for i, (label, value) in enumerate(girdi_rows, start=1):
    ws.cell(row=i, column=1, value=label).font = LABEL_FONT
    cell = ws.cell(row=i, column=2, value=value)
    cell.fill = INPUT_FILL
    cell.protection = Protection(locked=False)  # kullanıcı girebilir
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 34

# Hesap — formüller + kilitli sheet
ws = wb.create_sheet("Hesap")
ws["A1"] = "Kasa"
ws["B1"] = "=Girdi!B2"
ws["A2"] = "Gelir"
ws["B2"] = "=Girdi!B3"
ws["A3"] = "Gider"
ws["B3"] = "=Girdi!B4"
ws["A4"] = "Net Değişim"
ws["B4"] = "=B2-B3"
ws["A5"] = "Dönem Sonu"
ws["B5"] = "=B1+B4"
for row in range(1, 6):
    ws.cell(row=row, column=1).font = LABEL_FONT
ws.protection.sheet = True  # kilitli sheet kanıtı

# Çıktı — SUM + IF formülleri, kilitli sheet
ws = wb.create_sheet("Çıktı")
ws["A1"] = "Durum"
ws["B1"] = '=IF(Hesap!B5>0;"Kârda";"Dikkat")'
ws["A3"] = "Toplam Gelir"
ws["B3"] = "=SUM(Hesap!B2:B2)"
ws["A4"] = "Toplam Gider"
ws["B4"] = "=SUM(Hesap!B3:B3)"
ws["A5"] = "Fark"
ws["B5"] = "=B3-B4"
for row in range(1, 6):
    ws.cell(row=row, column=1).font = LABEL_FONT
ws.protection.sheet = True

wb.save(OUT)
print(f"OK -> {OUT} ({OUT.stat().st_size // 1024} KB)")
