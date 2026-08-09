#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Satış workbook'larından ürün sayfası ekran görüntüleri üretir.

Kullanım:
    python scripts/generate-screenshots.py <slug1> <slug2> ...

Her ürün için src/content/templates/<slug>.mdx frontmatter'ındaki screenshots
listesine uygun sayfa görüntüleri public/screenshots/<slug>-N.png olarak yazılır.

Yöntem: workbook kopyasında yalnız hedef sayfa görünür bırakılır, LibreOffice ile
PDF'e çevrilir, ilk sayfa ImageMagick ile 1200x750 PNG olarak dışa aktarılır.
"""
import argparse
import os
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile

import openpyxl
from openpyxl.utils import get_column_letter

KOK = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TESLIM = os.path.join(KOK, "delivery", "paid-products")
CIKTI = os.path.join(KOK, "public", "screenshots")
MDX_KONUM = os.path.join(KOK, "src", "content", "templates")

# Ürün başına ekran görüntüsünde kullanılacak sayfa sırası.
SAYFALAR = {
    "asiri-dusuk-teklif-savunma-robotu": ["RAKIP_TEKLIFLER", "KARAR", "PANO"],
    "ihaleye-kac-tl-teklif-vermeliyim": ["RAKIP_TEKLIFLER", "KARAR", "PANO"],
    "hakedis-fiyat-farki-hak-kaybi-cetveli": ["HAKEDIS_KALEMLERI", "KARAR", "PANO"],
    "yillara-sari-insaat-stopaj-nakit-akis-planlayici": ["HAKEDISLER", "KARAR", "PANO"],
    "taseron-hakedis-kesinti-mutabakati": ["HAKEDISLER", "KARAR", "PANO"],
    "kacirilan-sgk-tesvikleri-ve-gercek-iscilik-maliyeti-analizi": ["CALISAN_KARTI", "KARAR", "PANO"],
    "kidem-ihbar-yuku-ve-personel-cikarma-maliyeti-hesaplayici": ["CALISANLAR", "KARAR", "PANO"],
    "fazla-mesai-ve-isci-dava-riski-tespit-dosyasi": ["KAYITLAR", "KARAR", "PANO"],
    "asgari-ucret-zam-etkisi-fiyat-ayarlama-cetveli": ["CALISANLAR", "KARAR", "PANO"],
    "ithalat-depo-teslim-rafa-gelen-net-birim-maliyet": ["URUNLER", "KARAR", "PANO"],
}


def dolu_alan(ws):
    """Ekran görüntüsü için kompakt print_area üretir.

    Satır sınırı değer bazlıdır (formül taşması sayılmaz): son gerçek değerli
    satır + 3, üst sınır 40. Kolon sınırı formül dahildir (cached değerler
    görselde görünür): son dolu kolon + 2, üst sınır 18.
    """
    son_deger_satiri = 1
    son_kolon = 0
    for row in ws.iter_rows():
        dolu = [c for c in row if c.value not in (None, "")]
        degerli = [c for c in dolu
                   if not (isinstance(c.value, str) and c.value.startswith("="))]
        if degerli:
            son_deger_satiri = row[0].row
        for c in dolu:
            son_kolon = max(son_kolon, c.column)
    if son_kolon == 0:
        return None
    satir = min(son_deger_satiri + 3, 40)
    kolon = min(son_kolon + 2, 18)
    return "A1:%s%d" % (get_column_letter(kolon), satir)


def rel_haritasi(rels_xml):
    """workbook rels XML'inden rId → hedef dosya eşlemesi (attribute sırasından bağımsız)."""
    esleme = {}
    for r in re.finditer(r"<Relationship ([^>]*)/>", rels_xml):
        a = r.group(1)
        idd = re.search(r'Id="([^"]+)"', a)
        tgt = re.search(r'Target="([^"]+)"', a)
        if idd and tgt:
            esleme[idd.group(1)] = tgt.group(1)
    return esleme


def sayfa_görüntüle(kaynak, hedef_sayfa, cikti_png, soffice):
    with tempfile.TemporaryDirectory() as td:
        # 1) openpyxl yalnızca "workbook.xml" üretimi için kullanılır:
        #    yalnız hedef sayfa görünür yapılır. (Zip içi state düzenlemesi
        #    LibreOffice tarafından yok sayılıyor; openpyxl yazımı çalışıyor.)
        #    Cached değerler korunmadığı için paketin geri kalanı orijinal kalır.
        state_dosya = os.path.join(td, "state.xlsx")
        wb = openpyxl.load_workbook(kaynak)
        for ad in wb.sheetnames:
            wb[ad].sheet_state = "visible" if ad == hedef_sayfa else "hidden"
        alan = dolu_alan(wb[hedef_sayfa])
        wb.save(state_dosya)
        wb.close()
        if alan is None:
            return False
        with zipfile.ZipFile(state_dosya, "r") as z:
            yeni_wb = z.read("xl/workbook.xml")
            yeni_rels = z.read("xl/_rels/workbook.xml.rels")
        wbxml = yeni_wb.decode("utf-8")
        sayfa_rid = dict(re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wbxml))
        if hedef_sayfa not in sayfa_rid:
            return False
        # 2) Orijinal paketi kopyala; yalnız workbook.xml + rels değişir.
        kopya = os.path.join(td, "ekran.xlsx")
        shutil.copy(kaynak, kopya)
        with zipfile.ZipFile(kopya, "r") as z:
            girdiler = {n: z.read(n) for n in z.namelist()}
        girdiler["xl/workbook.xml"] = yeni_wb
        girdiler["xl/_rels/workbook.xml.rels"] = yeni_rels
        # 3) Hedef worksheet'e kompakt print alanı + tek sayfaya sığdırma.
        rel_map = rel_haritasi(yeni_rels.decode("utf-8"))
        hedef = rel_map[sayfa_rid[hedef_sayfa]].lstrip("/")
        if not hedef.startswith("xl/"):
            hedef = "xl/" + hedef
        wsxml = girdiler[hedef].decode("utf-8")
        wsxml = re.sub(r"<printArea[^/]*/>", "", wsxml)
        wsxml = re.sub(r"<pageSetup[^/]*/>", "", wsxml)
        wsxml = wsxml.replace(
            "</worksheet>",
            '<pageSetup fitToWidth="1" fitToHeight="1" orientation="landscape" paperSize="9" usePrinterDefaults="0" horizontalDpi="96" verticalDpi="96"/><printArea refs="%s"/></worksheet>'
            % alan)
        girdiler[hedef] = wsxml.encode("utf-8")
        with zipfile.ZipFile(kopya, "w", zipfile.ZIP_DEFLATED) as z:
            for n, d in girdiler.items():
                z.writestr(n, d)

        profil = os.path.join(td, "lo_profili")
        r = subprocess.run(
            [soffice, "--headless", "-env:UserInstallation=file://" + profil,
             "--convert-to", "pdf", "--outdir", td, kopya],
            capture_output=True, text=True, timeout=240)
        if r.returncode != 0:
            sys.stderr.write("LO hatası (%s): %s\n" % (hedef_sayfa, r.stderr[-400:]))
            return False
        pdf = os.path.join(td, "ekran.pdf")
        if not os.path.exists(pdf):
            sys.stderr.write("PDF üretilemedi: %s\n" % hedef_sayfa)
            return False
        png = os.path.join(td, "ekran.png")
        r = subprocess.run(
            ["gs", "-q", "-dNOPAUSE", "-dBATCH", "-sDEVICE=png16m", "-r96",
             "-dFirstPage=1", "-dLastPage=1", "-sOutputFile=" + png, pdf],
            capture_output=True, text=True, timeout=240)
        if r.returncode != 0 or not os.path.exists(png):
            sys.stderr.write("gs hatası (%s): %s\n" % (hedef_sayfa, r.stderr[-400:]))
            return False
        r = subprocess.run(
            ["magick", png, "-background", "white",
             "-resize", "1200x750^", "-gravity", "center", "-depth", "8",
             "-extent", "1200x750", cikti_png],
            capture_output=True, text=True, timeout=240)
        if r.returncode != 0 or not os.path.exists(cikti_png):
            sys.stderr.write("magick hatası (%s): %s\n" % (hedef_sayfa, r.stderr[-400:]))
            return False
        return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("sluglar", nargs="*")
    ap.add_argument("--soffice", default="soffice")
    a = ap.parse_args()

    slugs = a.sluglar or list(SAYFALAR)
    os.makedirs(CIKTI, exist_ok=True)
    for slug in slugs:
        if slug not in SAYFALAR:
            sys.stderr.write("Bilinmeyen slug: %s\n" % slug)
            continue
        kaynak = os.path.join(TESLIM, slug, "current.xlsx")
        if not os.path.exists(kaynak):
            sys.stderr.write("Teslim dosyası yok: %s\n" % kaynak)
            continue
        for i, sayfa in enumerate(SAYFALAR[slug], start=1):
            cikti = os.path.join(CIKTI, "%s-%d.png" % (slug, i))
            if sayfa_görüntüle(kaynak, sayfa, cikti, a.soffice):
                print("OK %s %s" % (slug, sayfa))
            else:
                print("HATA %s %s" % (slug, sayfa))


if __name__ == "__main__":
    main()
